#!/usr/bin/env python3
"""
check_xlsx.py - Workbook structure and style validator for tabtools QA

Validates cell contents, row-level formatting, merged cells, borders, fills,
and simple semantic text patterns using openpyxl.
"""

from __future__ import annotations

import os
import re
import sys
from collections import Counter
from typing import Iterable, List, Optional, Sequence, Tuple

from openpyxl.utils import column_index_from_string, get_column_letter


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    return str(value).strip()


def _is_option(token: str) -> bool:
    return token.startswith("--")


def _hex_to_rgb_string(hex_color: str) -> Optional[str]:
    if not hex_color:
        return None
    hex_color = hex_color.strip()
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    if len(hex_color) != 6:
        return None
    try:
        values = [str(int(hex_color[i : i + 2], 16)) for i in (0, 2, 4)]
    except ValueError:
        return None
    return " ".join(values)


def _cell_fill_rgb(cell) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "fill_type", None) != "solid":
        return None
    for color in (getattr(fill, "fgColor", None), getattr(fill, "start_color", None)):
        if color is None:
            continue
        rgb = getattr(color, "rgb", None)
        if rgb:
            parsed = _hex_to_rgb_string(rgb)
            if parsed is not None:
                return parsed
    return None


def _cell_has_border(cell) -> bool:
    border = getattr(cell, "border", None)
    if border is None:
        return False
    for side_name in ("top", "bottom", "left", "right"):
        side = getattr(border, side_name, None)
        if side is not None and getattr(side, "style", None):
            return True
    return False


def _cell_border_matches(cell, side_name: str, style: str) -> bool:
    border = getattr(cell, "border", None)
    if border is None:
        return False
    side = getattr(border, side_name, None)
    return side is not None and getattr(side, "style", None) == style


def _iter_nonempty_cells(ws) -> Iterable:
    for row in ws.iter_rows():
        for cell in row:
            if _stringify(cell.value) != "":
                yield cell


def _row_strings(ws, row: int) -> List[str]:
    return [_stringify(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)]


def _row_joined(ws, row: int) -> str:
    return " | ".join([value for value in _row_strings(ws, row) if value != ""])


def _all_text(ws) -> List[str]:
    values = []
    for cell in _iter_nonempty_cells(ws):
        values.append(_stringify(cell.value))
    return values


def _find_rows_containing(ws, text: str) -> List[int]:
    hits = []
    for row in range(1, ws.max_row + 1):
        if text in _row_joined(ws, row):
            hits.append(row)
    return hits


def _row_has_fill(ws, row: int, expected_rgb: Optional[str] = None) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        fill_rgb = _cell_fill_rgb(cell)
        if fill_rgb is None:
            continue
        if expected_rgb is None or fill_rgb == expected_rgb:
            return True
    return False


def _row_any_bold(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        if getattr(getattr(cell, "font", None), "bold", False):
            return True
    return False


def _row_all_bold(ws, row: int) -> bool:
    found = False
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        found = True
        if not getattr(getattr(cell, "font", None), "bold", False):
            return False
    return found


def _row_any_italic(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        if getattr(getattr(cell, "font", None), "italic", False):
            return True
    return False


def _dominant_font(ws) -> Tuple[str, Optional[float]]:
    counts: Counter = Counter()
    for cell in _iter_nonempty_cells(ws):
        font = getattr(cell, "font", None)
        if font is None:
            continue
        name = (font.name or "").strip()
        size = round(float(font.sz), 1) if font.sz is not None else None
        counts[(name, size)] += 1
    if not counts:
        return ("", None)
    return counts.most_common(1)[0][0]


def _sheet_has_pattern(ws, pattern_name: str) -> bool:
    texts = _all_text(ws)
    joined = "\n".join(texts)

    if pattern_name == "p-values":
        return any(
            re.search(r"(?i)(p-value|p\s*[=<]|<0\.\d+)", text)
            for text in texts
        )
    if pattern_name == "ci":
        return any(re.search(r"\([^)]*,\s*[^)]*\)", text) for text in texts) or "95% CI" in joined
    if pattern_name == "percentages":
        return any("%" in text for text in texts)
    if pattern_name == "mean-sd":
        return any(re.search(r"\d[^\n]*\([^,()%]*\)", text) for text in texts)
    if pattern_name == "n-equals":
        return any(re.search(r"\bN\s*=\s*\d+", text) for text in texts)
    if pattern_name == "sensitivity":
        return any("Sensitivity" in text for text in texts)
    return False


def _theme_expectations(name: str) -> Optional[Tuple[str, float, Optional[str]]]:
    theme = name.lower()
    if theme == "nejm":
        return ("Arial", 10.0, "medium")
    if theme == "lancet":
        return ("Arial", 9.0, "medium")
    if theme == "apa":
        return ("Times New Roman", 12.0, None)
    return None


def _column_width(ws, ref: str) -> float:
    ref = ref.strip().upper()
    try:
        col_idx = int(ref)
        letter = get_column_letter(col_idx)
    except ValueError:
        letter = get_column_letter(column_index_from_string(ref))
    width = ws.column_dimensions[letter].width
    return 8.43 if width is None else float(width)


def _parse_args(argv: Sequence[str]):
    if not argv:
        raise ValueError("Usage: check_xlsx.py <xlsx_file> [options]")

    xlsx_file = argv[0]
    sheet_name = None
    quiet = False
    result_file = None
    checks = []

    i = 1
    while i < len(argv):
        token = argv[i]

        if token == "--sheet":
            sheet_name = argv[i + 1]
            i += 2
        elif token == "--result-file":
            result_file = argv[i + 1]
            i += 2
        elif token == "--quiet":
            quiet = True
            i += 1
        elif token == "--min-rows":
            checks.append(("min_rows", int(argv[i + 1])))
            i += 2
        elif token == "--exact-rows":
            checks.append(("exact_rows", int(argv[i + 1])))
            i += 2
        elif token == "--min-cols":
            checks.append(("min_cols", int(argv[i + 1])))
            i += 2
        elif token == "--exact-cols":
            checks.append(("exact_cols", int(argv[i + 1])))
            i += 2
        elif token == "--cell":
            checks.append(("cell_exact", argv[i + 1], argv[i + 2]))
            i += 3
        elif token == "--cell-contains":
            checks.append(("cell_contains", argv[i + 1], argv[i + 2]))
            i += 3
        elif token == "--cell-regex":
            checks.append(("cell_regex", argv[i + 1], argv[i + 2]))
            i += 3
        elif token == "--cell-not-empty":
            i += 1
            while i < len(argv) and not _is_option(argv[i]):
                checks.append(("cell_not_empty", argv[i]))
                i += 1
        elif token == "--contains":
            checks.append(("contains", argv[i + 1]))
            i += 2
        elif token == "--row-contains":
            checks.append(("row_contains", int(argv[i + 1]), argv[i + 2]))
            i += 3
        elif token == "--header-row":
            row = int(argv[i + 1])
            labels = []
            i += 2
            while i < len(argv) and not _is_option(argv[i]):
                labels.append(argv[i])
                i += 1
            checks.append(("header_row", row, labels))
        elif token == "--has-borders":
            checks.append(("has_borders",))
            i += 1
        elif token == "--border-row":
            checks.append(("border_row", int(argv[i + 1]), argv[i + 2], argv[i + 3]))
            i += 4
        elif token == "--has-pattern":
            patterns = []
            i += 1
            while i < len(argv) and not _is_option(argv[i]):
                patterns.append(argv[i])
                i += 1
            for pattern_name in patterns:
                checks.append(("has_pattern", pattern_name))
        elif token == "--merged-row":
            checks.append(("merged_row", int(argv[i + 1])))
            i += 2
        elif token == "--col-width-at-least":
            checks.append(("col_width_at_least", argv[i + 1], float(argv[i + 2])))
            i += 3
        elif token == "--col-width-at-most":
            checks.append(("col_width_at_most", argv[i + 1], float(argv[i + 2])))
            i += 3
        elif token == "--min-merges":
            checks.append(("min_merges", int(argv[i + 1])))
            i += 2
        elif token == "--bold-row":
            checks.append(("bold_row", int(argv[i + 1])))
            i += 2
        elif token == "--bold-row-all":
            checks.append(("bold_row_all", int(argv[i + 1])))
            i += 2
        elif token == "--italic-row":
            checks.append(("italic_row", int(argv[i + 1])))
            i += 2
        elif token == "--italic-cell":
            checks.append(("italic_cell", argv[i + 1]))
            i += 2
        elif token == "--has-fill":
            checks.append(("has_fill", int(argv[i + 1])))
            i += 2
        elif token == "--fill-color":
            checks.append(("fill_color", int(argv[i + 1]), argv[i + 2]))
            i += 3
        elif token == "--font":
            checks.append(("font", argv[i + 1]))
            i += 2
        elif token == "--fontsize":
            checks.append(("fontsize", float(argv[i + 1])))
            i += 2
        elif token == "--theme":
            checks.append(("theme", argv[i + 1]))
            i += 2
        elif token == "--no-empty-cols":
            checks.append(("no_empty_cols",))
            i += 1
        elif token == "--row-bold-contains":
            checks.append(("row_bold_contains", argv[i + 1]))
            i += 2
        elif token == "--row-fill-contains":
            checks.append(("row_fill_contains", argv[i + 1], argv[i + 2]))
            i += 3
        else:
            raise ValueError(f"Unknown or incomplete option: {token}")

    return xlsx_file, sheet_name, quiet, result_file, checks


def _write_result(path: Optional[str], value: str) -> None:
    if path:
        with open(path, "w", encoding="utf-8") as handle:
            handle.write(value)


def main(argv: Sequence[str]) -> int:
    try:
        xlsx_file, sheet_name, quiet, result_file, checks = _parse_args(argv)
    except ValueError as exc:
        print(str(exc))
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("FAIL: openpyxl is required")
        _write_result(result_file, "FAIL")
        return 1

    if not os.path.exists(xlsx_file):
        print(f"FAIL: file not found: {xlsx_file}")
        _write_result(result_file, "FAIL")
        return 1

    try:
        workbook = load_workbook(xlsx_file, read_only=False, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive
        print(f"FAIL: could not open workbook: {exc}")
        _write_result(result_file, "FAIL")
        return 1

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            print(f"FAIL: sheet '{sheet_name}' not found")
            workbook.close()
            _write_result(result_file, "FAIL")
            return 1
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    failures: List[str] = []

    for check in checks:
        name = check[0]

        if name == "min_rows":
            if (ws.max_row or 0) < check[1]:
                failures.append(f"Expected at least {check[1]} rows, found {ws.max_row or 0}")

        elif name == "exact_rows":
            if (ws.max_row or 0) != check[1]:
                failures.append(f"Expected exactly {check[1]} rows, found {ws.max_row or 0}")

        elif name == "min_cols":
            if (ws.max_column or 0) < check[1]:
                failures.append(f"Expected at least {check[1]} columns, found {ws.max_column or 0}")

        elif name == "exact_cols":
            if (ws.max_column or 0) != check[1]:
                failures.append(f"Expected exactly {check[1]} columns, found {ws.max_column or 0}")

        elif name == "cell_exact":
            actual = _stringify(ws[check[1]].value)
            if actual != check[2]:
                failures.append(f"Cell {check[1]} expected '{check[2]}', found '{actual}'")

        elif name == "cell_contains":
            actual = _stringify(ws[check[1]].value)
            if check[2] not in actual:
                failures.append(f"Cell {check[1]} does not contain '{check[2]}'")

        elif name == "cell_regex":
            actual = _stringify(ws[check[1]].value)
            if re.search(check[2], actual) is None:
                failures.append(f"Cell {check[1]} does not match regex '{check[2]}'")

        elif name == "cell_not_empty":
            if _stringify(ws[check[1]].value) == "":
                failures.append(f"Cell {check[1]} is empty")

        elif name == "contains":
            if not any(check[1] in text for text in _all_text(ws)):
                failures.append(f"Workbook does not contain '{check[1]}'")

        elif name == "row_contains":
            row_text = _row_joined(ws, check[1])
            if check[2] not in row_text:
                failures.append(f"Row {check[1]} does not contain '{check[2]}'")

        elif name == "header_row":
            row_text = _row_joined(ws, check[1])
            for label in check[2]:
                if label not in row_text:
                    failures.append(f"Header row {check[1]} missing '{label}'")

        elif name == "has_borders":
            if not any(_cell_has_border(cell) for cell in _iter_nonempty_cells(ws)):
                failures.append("Workbook has no visible borders")

        elif name == "border_row":
            row, side_name, style = check[1], check[2], check[3]
            if not any(
                _cell_border_matches(ws.cell(row=row, column=col), side_name, style)
                for col in range(1, ws.max_column + 1)
            ):
                failures.append(f"Row {row} has no {style} {side_name} border")

        elif name == "has_pattern":
            if not _sheet_has_pattern(ws, check[1]):
                failures.append(f"Workbook missing pattern '{check[1]}'")

        elif name == "merged_row":
            if not any(rng.min_row <= check[1] <= rng.max_row for rng in ws.merged_cells.ranges):
                failures.append(f"No merged range touches row {check[1]}")

        elif name == "col_width_at_least":
            actual = _column_width(ws, check[1])
            if actual + 1e-9 < check[2]:
                failures.append(f"Column {check[1]} width {actual:.2f} is below minimum {check[2]:.2f}")

        elif name == "col_width_at_most":
            actual = _column_width(ws, check[1])
            if actual - 1e-9 > check[2]:
                failures.append(f"Column {check[1]} width {actual:.2f} exceeds maximum {check[2]:.2f}")

        elif name == "min_merges":
            if len(ws.merged_cells.ranges) < check[1]:
                failures.append(f"Expected at least {check[1]} merged ranges, found {len(ws.merged_cells.ranges)}")

        elif name == "bold_row":
            if not _row_any_bold(ws, check[1]):
                failures.append(f"Row {check[1]} is not bold")

        elif name == "bold_row_all":
            if not _row_all_bold(ws, check[1]):
                failures.append(f"Not all non-empty cells in row {check[1]} are bold")

        elif name == "italic_row":
            if not _row_any_italic(ws, check[1]):
                failures.append(f"Row {check[1]} is not italic")

        elif name == "italic_cell":
            if not getattr(getattr(ws[check[1]], "font", None), "italic", False):
                failures.append(f"Cell {check[1]} is not italic")

        elif name == "has_fill":
            if not _row_has_fill(ws, check[1]):
                failures.append(f"Row {check[1]} has no solid fill")

        elif name == "fill_color":
            if not _row_has_fill(ws, check[1], check[2]):
                failures.append(f"Row {check[1]} does not contain fill color '{check[2]}'")

        elif name == "font":
            dominant_name, _ = _dominant_font(ws)
            if dominant_name.lower() != check[1].lower():
                failures.append(f"Dominant font expected '{check[1]}', found '{dominant_name}'")

        elif name == "fontsize":
            _, dominant_size = _dominant_font(ws)
            if dominant_size is None or abs(dominant_size - check[1]) > 0.1:
                failures.append(f"Dominant font size expected {check[1]}, found {dominant_size}")

        elif name == "theme":
            expected = _theme_expectations(check[1])
            if expected is None:
                failures.append(f"Unknown theme expectation '{check[1]}'")
            else:
                exp_font, exp_size, exp_border = expected
                dominant_name, dominant_size = _dominant_font(ws)
                if dominant_name.lower() != exp_font.lower():
                    failures.append(f"Theme {check[1]} expected dominant font '{exp_font}', found '{dominant_name}'")
                if dominant_size is None or abs(dominant_size - exp_size) > 0.1:
                    failures.append(f"Theme {check[1]} expected dominant size {exp_size}, found {dominant_size}")
                if exp_border is not None and not any(
                    _cell_border_matches(cell, "bottom", exp_border) or _cell_border_matches(cell, "top", exp_border)
                    for cell in _iter_nonempty_cells(ws)
                ):
                    failures.append(f"Theme {check[1]} expected {exp_border} academic borders")

        elif name == "no_empty_cols":
            for col in range(1, ws.max_column + 1):
                if all(_stringify(ws.cell(row=row, column=col).value) == "" for row in range(1, ws.max_row + 1)):
                    failures.append(f"Column {col} is entirely empty")
                    break

        elif name == "row_bold_contains":
            rows = _find_rows_containing(ws, check[1])
            if not rows:
                failures.append(f"No row contains '{check[1]}'")
            elif not any(_row_any_bold(ws, row) for row in rows):
                failures.append(f"Rows containing '{check[1]}' are not bold")

        elif name == "row_fill_contains":
            rows = _find_rows_containing(ws, check[1])
            if not rows:
                failures.append(f"No row contains '{check[1]}'")
            elif not any(_row_has_fill(ws, row, check[2]) for row in rows):
                failures.append(f"Rows containing '{check[1]}' do not have fill '{check[2]}'")

    workbook.close()

    if failures:
        _write_result(result_file, "FAIL")
        if not quiet:
            for failure in failures:
                print(f"FAIL: {failure}")
        return 1

    _write_result(result_file, "PASS")
    if not quiet:
        print(f"PASS: {len(checks)} checks passed for {os.path.basename(xlsx_file)}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
