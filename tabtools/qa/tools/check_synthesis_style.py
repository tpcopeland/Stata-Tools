#!/usr/bin/env python3
"""
Workbook style assertions for the synthesized 4-reviewer audit (B1 and B2).

check_xlsx.py asserts that a border or an alignment IS present. These three
findings are all about a rule being present where it should not be, or absent
where it should be, so they need negative and paired assertions that the
canonical validator does not offer:

  B1  regtab emitted `op 9' (bottom rule) unguarded inside the stats() and
      addrow() loops, so every statistic row was underlined and a model-fit
      block rendered as a grid stapled to a booktabs table. Only the final row
      of the table may carry a bottom rule.

  B2  regtab set vertical=top on the label column alone, so C..N kept Excel's
      default bottom alignment and the stated intent -- the label's first line
      level with the single-line estimate cells -- failed on every row. Label
      and value cells in a body row must share a vertical alignment.

Usage:
    check_synthesis_style.py --b1 FILE --b1-sheet NAME
                             --b2 FILE --b2-sheet NAME
                             [--status-file PATH]

B1 and B2 take separate workbooks: B1 needs stats()/addrow() rows to exist,
while B2 must not see them, because those rows centre their own value cells by
design and would mask the alignment contract on ordinary coefficient rows.

Exit codes: 0 = all checks passed, 1 = a check failed, 2 = usage/read error.

Author: Timothy P Copeland, Karolinska Institutet
"""

import argparse
import sys

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - environment guard
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(2)


def _has_bottom(cell):
    b = cell.border.bottom
    return bool(b and b.style)


def _last_used_row(ws):
    last = 0
    for row in ws.iter_rows():
        if any(c.value not in (None, "") for c in row):
            last = row[0].row
    return last


def check_b1(path, sheet, failures):
    """No body row below the header rule may be underlined.

    Rows 1-3 are the title, the group header and the column header; the rule
    under the column header is the booktabs header rule and is correct. The
    defect was `op 9' emitted unguarded inside the stats() and addrow() loops,
    which underlined every one of those rows.
    """
    ws = load_workbook(path)[sheet]
    last = _last_used_row(ws)

    underlined = [
        r for r in range(4, last)
        if any(_has_bottom(ws.cell(row=r, column=c))
               for c in range(2, ws.max_column + 1))
    ]
    if underlined:
        failures.append(
            "B1: body rows carry a bottom rule: "
            + ", ".join(str(r) for r in underlined)
            + f" (only the closing rule on row {last} may)"
        )


def check_b2(path, sheet, failures):
    """Label and value cells in a body row must share a vertical alignment.

    Driven by a table with no stats()/addrow(), because those rows centre
    their own value cells by design; the defect was that on ordinary
    coefficient rows the label was top-aligned and C..N kept Excel's default
    bottom, so the stated intent failed on every single-line row.
    """
    ws = load_workbook(path)[sheet]
    last = _last_used_row(ws)

    mismatched = []
    for r in range(4, last + 1):
        label_v = ws.cell(row=r, column=2).alignment.vertical
        for c in range(3, ws.max_column + 1):
            if ws.cell(row=r, column=c).value in (None, ""):
                continue
            if ws.cell(row=r, column=c).alignment.vertical != label_v:
                mismatched.append(f"{r}:{c} ({label_v} vs "
                                  f"{ws.cell(row=r, column=c).alignment.vertical})")
                break
    if mismatched:
        failures.append(
            "B2: label and value cells disagree on vertical alignment at "
            + ", ".join(mismatched)
        )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--b1", required=True, help="regtab workbook WITH stats()/addrow()")
    ap.add_argument("--b1-sheet", required=True)
    ap.add_argument("--b2", required=True, help="regtab workbook WITHOUT stats()/addrow()")
    ap.add_argument("--b2-sheet", required=True)
    ap.add_argument("--status-file")
    args = ap.parse_args()

    failures = []
    try:
        check_b1(args.b1, args.b1_sheet, failures)
        check_b2(args.b2, args.b2_sheet, failures)
    except Exception as exc:  # noqa: BLE001 - surface the reason to Stata
        failures.append(f"error: {exc}")

    if failures:
        lines = ["FAIL synthesis style checks"] + failures
        rc = 1
    else:
        lines = ["PASS synthesis style checks (B1 B2)"]
        rc = 0

    out = "\n".join(lines)
    print(out)
    if args.status_file:
        with open(args.status_file, "w", encoding="utf-8") as fh:
            fh.write(out + "\n")
    return rc


if __name__ == "__main__":
    sys.exit(main())
