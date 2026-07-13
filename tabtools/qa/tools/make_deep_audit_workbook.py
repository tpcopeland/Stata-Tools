#!/usr/bin/env python3
"""Create deterministic Excel fixtures for tabtools deep-audit regressions."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


def build(path: Path, mode: str) -> None:
    workbook = Workbook()
    source = workbook.active
    source.title = "Source"
    source.append(["label", "value"])
    source.append(["new row", "sentinel-new"])

    target = workbook.create_sheet("Target")
    if mode == "dense":
        for row in range(1, 20006):
            target.cell(row=row, column=1, value=f"original-{row}")
        target.cell(row=20001, column=2, value="sentinel-20001")
        target.cell(row=20005, column=3, value="sentinel-20005")
    elif mode == "sparse":
        target["A1"] = "original-1"
        target["B500"] = "sentinel-500"
        target["C513"] = "sentinel-513"
    elif mode == "far-right":
        target["A1"] = "original-1"
        target["XFD700"] = "sentinel-xfd700"
    elif mode == "formatted-tail":
        target["A1"] = "original-1"
        target["B400"] = "sentinel-400"
        target["A900"].number_format = "0.00"
    elif mode == "numeric":
        target["A1"] = 1001
        target["B777"] = 42.5
    elif mode == "case":
        target["A1"] = "original-1"
        target["B2"] = "sentinel-case"
    else:
        raise ValueError(f"unknown fixture mode: {mode}")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _target_cells(path: Path) -> dict[str, dict[str, object]]:
    """Return a stable payload for cells instantiated before stacktab appends.

    Accessing ``_cells`` is intentional here: iterating to XFD700 would create
    more than eleven million empty cells and would obscure the far-right-edge
    regression this fixture is designed to test.
    """

    workbook = load_workbook(path, data_only=False)
    target = workbook["Target"]
    payload: dict[str, dict[str, object]] = {}
    for cell in sorted(target._cells.values(), key=lambda item: item.coordinate):
        payload[cell.coordinate] = {
            "value": cell.value,
            "data_type": cell.data_type,
            "number_format": cell.number_format,
            "style_id": cell.style_id,
        }
    workbook.close()
    return payload


def snapshot(path: Path, signature_path: Path) -> None:
    payload = _target_cells(path)
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    signature_path.write_text(
        json.dumps(
            {
                "cells": payload,
                "sha256": hashlib.sha256(encoded).hexdigest(),
            },
            ensure_ascii=False,
            sort_keys=True,
        ),
        encoding="utf-8",
    )


def verify(path: Path, signature_path: Path) -> None:
    expected = json.loads(signature_path.read_text(encoding="utf-8"))
    current_all = _target_cells(path)
    current = {coordinate: current_all.get(coordinate) for coordinate in expected["cells"]}
    encoded = json.dumps(current, ensure_ascii=False, sort_keys=True).encode("utf-8")
    current_hash = hashlib.sha256(encoded).hexdigest()
    if current != expected["cells"] or current_hash != expected["sha256"]:
        raise SystemExit("pre-existing Target cells changed after append")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("path", type=Path)
    parser.add_argument(
        "mode",
        choices=("dense", "sparse", "far-right", "formatted-tail", "numeric", "case"),
    )
    parser.add_argument("--snapshot", type=Path)
    parser.add_argument("--verify", type=Path)
    args = parser.parse_args()
    if args.snapshot and args.verify:
        parser.error("choose only one of --snapshot or --verify")
    if args.snapshot:
        snapshot(args.path, args.snapshot)
    elif args.verify:
        verify(args.path, args.verify)
    else:
        build(args.path, args.mode)


if __name__ == "__main__":
    main()
