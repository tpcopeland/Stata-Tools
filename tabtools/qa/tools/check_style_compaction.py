#!/usr/bin/env python3
"""
check_style_compaction.py - independent oracle for xlsx style-pool compaction.

Two jobs, neither of which trusts the Stata-side rewriter:

  --compare A B   every cell of every sheet in A must resolve, through
                  openpyxl, to the same value, font, fill, border and
                  alignment as in B, along with row heights, column widths
                  and merged ranges.  This is what makes collapsing the
                  style pools a format-preserving transformation rather
                  than a hopeful one.

  --counts F      report the size of each style pool declared in
                  xl/styles.xml, read straight from the archive.
"""

from __future__ import annotations

import argparse
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook


def _color(color) -> str:
    if color is None:
        return ""
    parts = []
    for attr in ("type", "rgb", "tint"):
        try:
            value = getattr(color, attr, "")
        except Exception:
            value = ""
        parts.append("" if value is None else str(value))
    return "/".join(parts)


def cell_style(cell) -> tuple:
    font, fill, align, border = cell.font, cell.fill, cell.alignment, cell.border
    return (
        font.name, font.sz, bool(font.b), bool(font.i), bool(font.u),
        _color(font.color),
        fill.patternType, _color(fill.fgColor), _color(fill.bgColor),
        align.horizontal, align.vertical, bool(align.wrapText), align.indent,
        border.top.style, border.bottom.style,
        border.left.style, border.right.style,
        _color(border.top.color), _color(border.bottom.color),
        _color(border.left.color), _color(border.right.color),
        cell.number_format,
    )


def signature(path: Path) -> dict:
    wb = load_workbook(path, read_only=False, data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and cell.style_id == 0:
                    continue
                cells[cell.coordinate] = (cell.value, cell_style(cell))
        out[name] = {
            "cells": cells,
            "row_heights": {
                r: d.height for r, d in ws.row_dimensions.items() if d.height is not None
            },
            "col_widths": {
                c: d.width for c, d in ws.column_dimensions.items() if d.width is not None
            },
            "merges": sorted(str(rng) for rng in ws.merged_cells.ranges),
        }
    wb.close()
    return out


def compare(a: Path, b: Path, only: str = None) -> list:
    sa, sb = signature(a), signature(b)
    failures = []
    if only:
        for side, sig in (("A", sa), ("B", sb)):
            if only not in sig:
                return ["sheet %s not in %s" % (only, side)]
        names = [only]
    else:
        if sorted(sa) != sorted(sb):
            return ["sheet names differ: %s vs %s" % (sorted(sa), sorted(sb))]
        names = list(sa)
    for name in names:
        for key in ("row_heights", "col_widths", "merges"):
            if sa[name][key] != sb[name][key]:
                failures.append("%s: %s differ" % (name, key))
        ca, cb = sa[name]["cells"], sb[name]["cells"]
        if set(ca) != set(cb):
            missing = sorted(set(ca) ^ set(cb))[:5]
            failures.append("%s: styled cell set differs (e.g. %s)" % (name, missing))
        for ref in sorted(set(ca) & set(cb)):
            if ca[ref] != cb[ref]:
                failures.append("%s!%s: %r vs %r" % (name, ref, ca[ref], cb[ref]))
                if len(failures) > 20:
                    return failures
    return failures


def counts(path: Path) -> dict:
    with zipfile.ZipFile(path) as z:
        styles = z.read("xl/styles.xml").decode("utf-8", "replace")
    out = {}
    for tag in ("fonts", "fills", "borders", "cellXfs"):
        m = re.search(r'<%s count="(\d+)"' % tag, styles)
        out[tag] = int(m.group(1)) if m else 0
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--compare", nargs=2, metavar=("A", "B"))
    parser.add_argument("--sheet", help="restrict --compare to one sheet")
    parser.add_argument("--counts", metavar="FILE")
    parser.add_argument("--result-file")
    args = parser.parse_args()

    lines = []
    failures = []

    if args.counts:
        try:
            for tag, n in counts(Path(args.counts)).items():
                lines.append("%s=%d" % (tag, n))
        except Exception as exc:
            failures.append("counts: %s" % exc)

    if args.compare:
        try:
            failures.extend(
                compare(Path(args.compare[0]), Path(args.compare[1]), args.sheet)
            )
        except Exception as exc:
            failures.append("compare: %s" % exc)

    if args.result_file:
        text = "FAIL" if failures else "PASS"
        if lines:
            text = text + "\n" + "\n".join(lines)
        Path(args.result_file).write_text(text + "\n", encoding="utf-8")

    for line in lines:
        print(line)
    for failure in failures:
        print("FAIL: %s" % failure, file=sys.stderr)
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
