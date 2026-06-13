#!/usr/bin/env python3
"""Small workbook checks for stacktab QA."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


# Path of the optional --result-file. Stata's `shell` command does not propagate
# child exit codes to _rc, so the verdict is recorded here for the .do to assert.
_RESULT_FILE: Path | None = None


def _write_result(verdict: str, details: str = "") -> None:
    if _RESULT_FILE is None:
        return
    line = verdict if not details else f"{verdict}\t{details}"
    _RESULT_FILE.write_text(line)


def fail(message: str) -> None:
    print(f"FAIL: {message}", file=sys.stderr)
    _write_result("FAIL", message)
    raise SystemExit(1)


def close_enough(actual: float | None, expected: float, tolerance: float = 0.5) -> bool:
    if actual is None:
        return False
    return abs(float(actual) - expected) <= tolerance


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("sheet")
    parser.add_argument("--row-height", nargs=2, action="append", metavar=("ROW", "HEIGHT"))
    parser.add_argument("--col-width", nargs=2, action="append", metavar=("COL", "WIDTH"))
    parser.add_argument("--border-range", nargs=2, action="append", metavar=("START", "END"))
    parser.add_argument("--outer-border", nargs=2, action="append", metavar=("START", "END"))
    parser.add_argument("--cell", nargs=2, action="append", metavar=("CELL", "VALUE"))
    parser.add_argument("--blank", action="append", metavar="CELL")
    parser.add_argument("--merged", action="append", metavar="RANGE")
    parser.add_argument("--bold", action="append", metavar="CELL")
    parser.add_argument("--italic", action="append", metavar="CELL")
    parser.add_argument(
        "--result-file",
        metavar="PATH",
        help="Write PASS/FAIL verdict to this file for Stata integration "
        "(exit codes are not reliably propagated by Stata's shell command).",
    )
    args = parser.parse_args()

    global _RESULT_FILE
    if args.result_file:
        _RESULT_FILE = Path(args.result_file)

    path = Path(args.workbook)
    if not path.exists():
        fail(f"workbook not found: {path}")

    wb = load_workbook(path)
    if args.sheet not in wb.sheetnames:
        fail(f"sheet not found: {args.sheet}")
    ws = wb[args.sheet]

    for cell, expected in args.cell or []:
        actual = ws[cell].value
        if actual != expected:
            fail(f"{cell} value {actual!r} != {expected!r}")

    for cell in args.blank or []:
        actual = ws[cell].value
        if actual not in (None, ""):
            fail(f"{cell} value {actual!r} is not blank")

    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    for expected_range in args.merged or []:
        if expected_range not in merged_ranges:
            fail(f"merged range {expected_range!r} not found")

    for cell in args.bold or []:
        if not ws[cell].font.bold:
            fail(f"{cell} is not bold")

    for cell in args.italic or []:
        if not ws[cell].font.italic:
            fail(f"{cell} is not italic")

    for row, height in args.row_height or []:
        row_index = int(row)
        expected = float(height)
        actual = ws.row_dimensions[row_index].height
        if not close_enough(actual, expected):
            fail(f"row {row_index} height {actual!r} != {expected}")

    for col, width in args.col_width or []:
        expected = float(width)
        actual = ws.column_dimensions[col].width
        # A column width set via Mata's xl() is read back by openpyxl with a
        # constant Excel padding offset of ~0.711 chars (5px / 7px-per-char),
        # e.g. a requested width of 24 reads back as 24.7109375. Allow a
        # tolerance wide enough to absorb that read-back artifact.
        if not close_enough(actual, expected, tolerance=1.0):
            fail(f"column {col} width {actual!r} != {expected} (+/-1.0)")

    for start, end in args.border_range or []:
        cells = ws[f"{start}:{end}"]
        for row in cells:
            for cell in row:
                styles = [
                    cell.border.top.style,
                    cell.border.bottom.style,
                    cell.border.left.style,
                    cell.border.right.style,
                ]
                if not any(style == "thin" for style in styles):
                    fail(f"{cell.coordinate} has no thin border")

    for start, end in args.outer_border or []:
        cells = ws[f"{start}:{end}"]
        top = cells[0]
        bottom = cells[-1]
        left = [row[0] for row in cells]
        right = [row[-1] for row in cells]

        for cell in top:
            if cell.border.top.style != "thin":
                fail(f"{cell.coordinate} has no thin top border")
        for cell in bottom:
            if cell.border.bottom.style != "thin":
                fail(f"{cell.coordinate} has no thin bottom border")
        for cell in left:
            if cell.border.left.style != "thin":
                fail(f"{cell.coordinate} has no thin left border")
        for cell in right:
            if cell.border.right.style != "thin":
                fail(f"{cell.coordinate} has no thin right border")

    print("PASS: workbook checks passed")
    _write_result("PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
