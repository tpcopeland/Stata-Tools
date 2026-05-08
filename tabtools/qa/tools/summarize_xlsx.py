#!/usr/bin/env python3
"""
Write a compact workbook-summary TSV for refactor baselines.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def cell_text(cell) -> str:
    value = cell.value
    return "" if value is None else str(value)


def cell_fill_rgb(cell) -> str:
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "fill_type", None) != "solid":
        return ""
    color = getattr(fill, "fgColor", None)
    rgb = getattr(color, "rgb", None) if color is not None else None
    if not rgb:
        return ""
    return rgb[-6:].upper()


def cell_has_border(cell) -> bool:
    border = getattr(cell, "border", None)
    if border is None:
        return False
    for side_name in ("top", "bottom", "left", "right"):
        side = getattr(border, side_name, None)
        if side is not None and getattr(side, "style", None):
            return True
    return False


def dominant_font(ws) -> tuple[str, str]:
    counts: Counter[tuple[str, str]] = Counter()
    for row in ws.iter_rows():
        for cell in row:
            if cell_text(cell) == "":
                continue
            font = getattr(cell, "font", None)
            if font is None:
                continue
            name = (font.name or "").strip()
            size = ""
            if font.sz is not None:
                size = str(round(float(font.sz), 1))
            counts[(name, size)] += 1
    if not counts:
        return ("", "")
    return counts.most_common(1)[0][0]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Summarize a workbook sheet to TSV")
    parser.add_argument("xlsx")
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--result-file", required=True)
    parser.add_argument("--expect-file")
    parser.add_argument("--compare-columns", nargs="+")
    return parser.parse_args()


def write_status(path: Path, status: str, message: str) -> None:
    with path.open("w", newline="") as fh:
        writer = csv.writer(fh, delimiter="\t", lineterminator="\n")
        writer.writerow(["status", "message"])
        writer.writerow([status, message])


def summarize_workbook(xlsx: str, sheet: str) -> dict[str, object]:
    wb = load_workbook(xlsx)
    ws = wb[sheet]

    n_bold = 0
    n_filled = 0
    n_border = 0
    fill_palette: Counter[str] = Counter()
    nonempty_payload: list[str] = []

    for row in ws.iter_rows():
        for cell in row:
            text = cell_text(cell)
            if text == "":
                continue
            nonempty_payload.append(f"{cell.coordinate}={text}")
            font = getattr(cell, "font", None)
            if font is not None and getattr(font, "bold", False):
                n_bold += 1
            fill_rgb = cell_fill_rgb(cell)
            if fill_rgb:
                n_filled += 1
                fill_palette[fill_rgb] += 1
            if cell_has_border(cell):
                n_border += 1

    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
    font_name, font_size = dominant_font(ws)
    top_fills = ";".join(color for color, _ in fill_palette.most_common(5))
    digest_input = "\n".join(nonempty_payload).encode("utf-8")

    return {
        "status": "PASS",
        "sheet": ws.title,
        "title": cell_text(ws["A1"]),
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "n_merges": len(merged_ranges),
        "merged_ranges": ";".join(merged_ranges),
        "n_bold_cells": n_bold,
        "n_filled_cells": n_filled,
        "n_border_cells": n_border,
        "dominant_font": font_name,
        "dominant_font_size": font_size,
        "fill_palette": top_fills,
        "nonempty_text_count": len(nonempty_payload),
        "content_digest": hashlib.sha256(digest_input).hexdigest(),
    }


def write_summary(path: Path, summary: dict[str, object]) -> None:
    fields = [
        "status",
        "sheet",
        "title",
        "max_row",
        "max_col",
        "n_merges",
        "merged_ranges",
        "n_bold_cells",
        "n_filled_cells",
        "n_border_cells",
        "dominant_font",
        "dominant_font_size",
        "fill_palette",
        "nonempty_text_count",
        "content_digest",
    ]
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, delimiter="\t", fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerow({field: summary[field] for field in fields})


def read_expected(path: Path) -> dict[str, str]:
    with path.open(newline="") as fh:
        reader = csv.DictReader(fh, delimiter="\t")
        rows = list(reader)
    if len(rows) != 1:
        raise ValueError(f"{path} must contain exactly one summary row")
    return rows[0]


def main() -> int:
    args = parse_args()
    result_file = Path(args.result_file)

    if load_workbook is None:
        write_status(result_file, "FAIL", "openpyxl not installed")
        return 1

    try:
        summary = summarize_workbook(args.xlsx, args.sheet)
    except Exception as exc:  # pragma: no cover - defensive
        write_status(result_file, "FAIL", str(exc))
        return 1

    write_summary(result_file, summary)
    if args.expect_file:
        try:
            expected = read_expected(Path(args.expect_file))
        except Exception as exc:
            print(f"FAIL: could not read expected summary: {exc}", file=sys.stderr)
            return 1
        compare_columns = args.compare_columns or sorted(expected.keys() & summary.keys())
        failures = []
        for column in compare_columns:
            actual = str(summary.get(column, ""))
            wanted = str(expected.get(column, ""))
            if actual != wanted:
                failures.append(f"{column}: expected {wanted!r}, found {actual!r}")
        if failures:
            for failure in failures:
                print(f"FAIL: {failure}", file=sys.stderr)
            return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
