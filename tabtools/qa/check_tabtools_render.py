#!/usr/bin/env python3
"""
Focused Excel rendering assertions for tabtools regression QA.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize_rgb(rgb: str) -> str:
    if not rgb:
        return ""
    rgb = rgb.strip()
    compact = rgb.replace(" ", "")
    if len(compact) == 6:
        return compact.upper()
    parts = rgb.split(",") if "," in rgb else rgb.split()
    if len(parts) == 3:
        return "".join(f"{int(part):02X}" for part in parts).upper()
    return compact[-6:].upper()


def cell_text(ws, ref: str) -> str:
    value = ws[ref].value
    return "" if value is None else str(value)


def cell_is_bold(ws, ref: str) -> bool:
    font = ws[ref].font
    return bool(font and font.bold)


def row_text(ws, row: int) -> str:
    values = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        values.append(str(value))
    return " | ".join(values)


def rows_containing(ws, text: str) -> list[int]:
    hits: list[int] = []
    for row in range(1, ws.max_row + 1):
        if text in row_text(ws, row):
            hits.append(row)
    return hits


def row_has_bold(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            continue
        font = cell.font
        if font and font.bold:
            return True
    return False


def row_has_fill(ws, row: int, rgb: str) -> bool:
    target = normalize_rgb(rgb)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            continue
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            continue
        fill_rgb = ""
        if fill.fgColor is not None:
            if fill.fgColor.type == "rgb" and fill.fgColor.rgb:
                fill_rgb = fill.fgColor.rgb[-6:]
            elif fill.fgColor.type == "indexed" and fill.fgColor.indexed is not None:
                fill_rgb = str(fill.fgColor.indexed)
        if fill_rgb.upper() == target:
            return True
    return False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="tabtools Excel rendering checks")
    parser.add_argument("xlsx")
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--cell-contains", nargs=2, action="append", default=[], metavar=("REF", "TEXT"))
    parser.add_argument("--cell-bold", action="append", default=[], metavar="REF")
    parser.add_argument("--row-fill", nargs=2, action="append", default=[], metavar=("ROW", "RGB"))
    parser.add_argument("--row-contains-bold", action="append", default=[], metavar="TEXT")
    parser.add_argument("--row-contains-fill", nargs=2, action="append", default=[], metavar=("TEXT", "RGB"))
    parser.add_argument("--result-file", required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    wb = load_workbook(args.xlsx)
    ws = wb[args.sheet]

    failures: list[str] = []

    for ref, text in args.cell_contains:
        if text not in cell_text(ws, ref):
            failures.append(f"{ref} missing substring: {text}")

    for ref in args.cell_bold:
        if not cell_is_bold(ws, ref):
            failures.append(f"{ref} is not bold")

    for row, rgb in args.row_fill:
        if not row_has_fill(ws, int(row), rgb):
            failures.append(f"row {row} missing fill {rgb}")

    for text in args.row_contains_bold:
        hits = rows_containing(ws, text)
        if not hits:
            failures.append(f"no row contains text: {text}")
        elif not any(row_has_bold(ws, row) for row in hits):
            failures.append(f"row containing '{text}' is not bold")

    for text, rgb in args.row_contains_fill:
        hits = rows_containing(ws, text)
        if not hits:
            failures.append(f"no row contains text: {text}")
        elif not any(row_has_fill(ws, row, rgb) for row in hits):
            failures.append(f"row containing '{text}' missing fill {rgb}")

    Path(args.result_file).write_text("PASS" if not failures else "FAIL")
    if failures:
        for failure in failures:
            print(failure, file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
