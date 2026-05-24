#!/usr/bin/env python3
"""
Compare tabtools shared style-engine workbook fixtures.

The comparator checks rendered cell values plus workbook presentation surfaces
that a shared Excel style engine can accidentally change: styles, merged
ranges, row heights, column widths, borders, and fills.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def color_signature(color: Any) -> tuple:
    if color is None:
        return ("", "", "")
    return (
        str(getattr(color, "type", "") or ""),
        str(getattr(color, "rgb", "") or ""),
        str(getattr(color, "indexed", "") or ""),
    )


def side_signature(side: Any) -> tuple:
    if side is None:
        return ("", ("", "", ""))
    return (str(getattr(side, "style", "") or ""), color_signature(getattr(side, "color", None)))


def cell_signature(cell: Any) -> dict:
    font = cell.font
    fill = cell.fill
    align = cell.alignment
    border = cell.border
    number_format = "" if cell.number_format is None else str(cell.number_format)
    return {
        "value": "" if cell.value is None else cell.value,
        "font": (
            font.name,
            float(font.sz) if font.sz is not None else None,
            bool(font.bold),
            bool(font.italic),
            color_signature(font.color),
        ),
        "fill": (
            fill.fill_type,
            color_signature(fill.fgColor),
            color_signature(fill.bgColor),
        ),
        "alignment": (
            align.horizontal,
            align.vertical,
            bool(align.wrap_text),
            int(align.text_rotation or 0),
        ),
        "border": (
            side_signature(border.top),
            side_signature(border.bottom),
            side_signature(border.left),
            side_signature(border.right),
        ),
        "number_format": number_format,
    }


def workbook_signature(path: Path, sheet: str) -> dict:
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"{path.name}: sheet not found: {sheet}")
        ws = wb[sheet]
        max_row = max(ws.max_row, 1)
        max_col = max(ws.max_column, 1)
        cells = {}
        has_border = False
        has_fill = False
        nonempty = 0

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                sig = cell_signature(cell)
                ref = f"{get_column_letter(col)}{row}"
                cells[ref] = sig
                if sig["value"] != "":
                    nonempty += 1
                if any(side[0] for side in sig["border"]):
                    has_border = True
                if sig["fill"][0] not in ("", None):
                    has_fill = True

        return {
            "dimensions": (max_row, max_col),
            "cells": cells,
            "merges": sorted(str(rng) for rng in ws.merged_cells.ranges),
            "row_heights": {
                str(idx): ws.row_dimensions[idx].height
                for idx in range(1, max_row + 1)
                if ws.row_dimensions[idx].height is not None
            },
            "col_widths": {
                get_column_letter(idx): ws.column_dimensions[get_column_letter(idx)].width
                for idx in range(1, max_col + 1)
                if ws.column_dimensions[get_column_letter(idx)].width is not None
            },
            "nonempty": nonempty,
            "has_border": has_border,
            "has_fill": has_fill,
        }
    finally:
        wb.close()


def compare_workbooks(before: Path, after: Path, sheet: str) -> tuple[bool, list[str], dict]:
    before_sig = workbook_signature(before, sheet)
    after_sig = workbook_signature(after, sheet)
    failures: list[str] = []

    if before_sig["nonempty"] == 0:
        failures.append("before workbook has no populated cells")
    if after_sig["nonempty"] == 0:
        failures.append("after workbook has no populated cells")
    if not before_sig["has_border"]:
        failures.append("before workbook has no styled borders to protect")
    if not after_sig["has_border"]:
        failures.append("after workbook has no styled borders")
    if not before_sig["has_fill"]:
        failures.append("before workbook has no fills to protect")
    if not after_sig["has_fill"]:
        failures.append("after workbook has no fills")

    for key in ("dimensions", "cells", "merges", "row_heights", "col_widths"):
        if before_sig[key] != after_sig[key]:
            failures.append(f"{key} differ")

    summary = {
        "before": str(before),
        "after": str(after),
        "sheet": sheet,
        "nonempty": after_sig["nonempty"],
        "merges": after_sig["merges"],
        "has_border": after_sig["has_border"],
        "has_fill": after_sig["has_fill"],
        "failures": failures,
    }
    return (len(failures) == 0, failures, summary)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("before")
    parser.add_argument("after")
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--result-file")
    parser.add_argument("--report-file")
    args = parser.parse_args()

    try:
        ok, failures, summary = compare_workbooks(
            Path(args.before), Path(args.after), args.sheet
        )
    except Exception as exc:
        ok = False
        failures = [str(exc)]
        summary = {
            "before": args.before,
            "after": args.after,
            "sheet": args.sheet,
            "failures": failures,
        }

    if args.report_file:
        Path(args.report_file).write_text(
            json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8"
        )
    if args.result_file:
        Path(args.result_file).write_text("PASS" if ok else "FAIL", encoding="utf-8")

    if ok:
        print("PASS: workbook values/styles/merges/borders/fills match")
        return 0
    for failure in failures:
        print(f"FAIL: {failure}", file=sys.stderr)
    return 1


if __name__ == "__main__":
    sys.exit(main())
