#!/usr/bin/env python3
"""
Compare table1_tc before/after Excel artifacts for value and style parity.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def color_rgb(color: Any) -> str:
    if color is None:
        return ""
    rgb = getattr(color, "rgb", None)
    if rgb:
        return str(rgb)
    indexed = getattr(color, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"
    return ""


def side_style(side: Any) -> tuple[str, str]:
    if side is None:
        return ("", "")
    return (str(getattr(side, "style", "") or ""), color_rgb(getattr(side, "color", None)))


def cell_signature(cell: Any) -> tuple:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    return (
        cell.value,
        font.name,
        float(font.sz) if font.sz is not None else None,
        bool(font.bold),
        bool(font.italic),
        alignment.horizontal,
        alignment.vertical,
        bool(alignment.wrap_text),
        fill.fill_type,
        color_rgb(fill.fgColor),
        side_style(border.top),
        side_style(border.bottom),
        side_style(border.left),
        side_style(border.right),
        cell.number_format,
    )


def sheet_signature(path: Path, sheet: str) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=True)
    if sheet not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"sheet not found in {path.name}: {sheet}")
    ws = workbook[sheet]
    sig = {
        "size": (ws.max_row, ws.max_column),
        "merges": sorted(str(rng) for rng in ws.merged_cells.ranges),
        "rows": [
            [cell_signature(ws.cell(row=row, column=col)) for col in range(1, ws.max_column + 1)]
            for row in range(1, ws.max_row + 1)
        ],
        "row_heights": {
            row: ws.row_dimensions[row].height for row in range(1, ws.max_row + 1)
        },
        "col_widths": {
            col: ws.column_dimensions[get_column_letter(col)].width
            for col in range(1, ws.max_column + 1)
        },
    }
    workbook.close()
    return sig


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("before")
    parser.add_argument("after")
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--result-file")
    args = parser.parse_args()

    failures = []
    try:
        before = sheet_signature(Path(args.before), args.sheet)
        after = sheet_signature(Path(args.after), args.sheet)
    except Exception as exc:
        failures.append(str(exc))
    else:
        for key in ("size", "merges", "rows", "row_heights", "col_widths"):
            if before[key] != after[key]:
                failures.append(f"{key} differ")

    if args.result_file:
        Path(args.result_file).write_text("FAIL" if failures else "PASS", encoding="utf-8")

    if failures:
        for failure in failures:
            print(f"FAIL: {failure}", file=sys.stderr)
        return 1

    print("PASS: table1_tc before/after workbook signatures match")
    return 0


if __name__ == "__main__":
    sys.exit(main())
