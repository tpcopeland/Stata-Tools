#!/usr/bin/env python3
"""
style_engine_compare.py - compare legacy and shared-style Excel workbooks.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def color_rgb(color) -> str:
    if color is None:
        return ""
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return ""
    return str(rgb)


def cell_style(cell) -> tuple:
    font = cell.font
    fill = cell.fill
    align = cell.alignment
    border = cell.border
    return (
        font.name,
        float(font.sz) if font.sz is not None else None,
        bool(font.bold),
        bool(font.italic),
        align.horizontal,
        align.vertical,
        bool(align.wrap_text),
        fill.fill_type,
        color_rgb(fill.fgColor),
        border.top.style,
        border.bottom.style,
        border.left.style,
        border.right.style,
    )


def workbook_signature(path: Path, sheet: str) -> dict:
    wb = load_workbook(path, read_only=False, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"sheet not found: {sheet}")
    ws = wb[sheet]
    refs = ("A1", "A2", "B2", "E2", "B3", "E3", "A4", "D4", "A5", "A6")
    signature = {
        "values": {ref: ws[ref].value for ref in refs},
        "styles": {ref: cell_style(ws[ref]) for ref in refs},
        "row_heights": {row: ws.row_dimensions[row].height for row in (1, 2, 6)},
        "col_widths": {
            col: ws.column_dimensions[col].width
            for col in ("A", "B", "C", "D", "E")
        },
        "merges": sorted(str(rng) for rng in ws.merged_cells.ranges),
    }
    wb.close()
    return signature


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("legacy")
    parser.add_argument("engine")
    parser.add_argument("--sheet", default="Style")
    parser.add_argument("--result-file")
    args = parser.parse_args()

    failures = []
    try:
        legacy = workbook_signature(Path(args.legacy), args.sheet)
        engine = workbook_signature(Path(args.engine), args.sheet)
    except Exception as exc:
        failures.append(str(exc))
    else:
        for key in ("values", "styles", "row_heights", "col_widths", "merges"):
            if legacy[key] != engine[key]:
                failures.append(f"{key} differ")

    if args.result_file:
        Path(args.result_file).write_text("FAIL" if failures else "PASS", encoding="utf-8")

    if failures:
        for failure in failures:
            print(f"FAIL: {failure}", file=sys.stderr)
        return 1

    print("PASS: style signatures match")
    return 0


if __name__ == "__main__":
    sys.exit(main())
