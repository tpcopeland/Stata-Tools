#!/usr/bin/env python3
"""Validate styled iivw reporting workbooks."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SPECS = {
    "balance": {
        "row2": {
            3: "Means",
            6: "Balance",
            9: "Counts",
        },
        "row3": [
            "",
            "Covariate",
            "Unweighted mean",
            "Weighted mean",
            "Unweighted SD",
            "SMD",
            "|SMD|",
            "Modeled",
            "N",
            "Missing",
        ],
        "merges": ("C2:E2", "F2:H2", "I2:J2"),
        "probe": "C4",
    },
    "diagnostics": {
        "row2": {
            3: "Model estimates",
        },
        "row3": [
            "",
            "Quantity",
            "Estimate",
            "SE",
            "95% CI",
        ],
        "merges": ("C2:E2",),
        "probe": "C4",
        # Single-value diagnostic block: a bold "Diagnostic values" divider
        # row bracketed by full-width horizontal rules (not merged), with each
        # value below it sitting plainly in column C (not merged).
        "divider": {"row": 7, "label": "Diagnostic values"},
    },
    "exogeneity": {
        "probe": "C4",
    },
}


def fail(message: str) -> None:
    print(f"FAIL: {message}", file=sys.stderr)
    raise SystemExit(1)


def check(
    path: Path,
    sheet: str,
    mode: str,
    expected_rows: int | None,
    marker: Path | None,
) -> None:
    if mode not in SPECS:
        fail("mode must be balance, diagnostics, or exogeneity")
    if not path.exists():
        fail(f"workbook not found: {path}")

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        fail(f"sheet not found: {sheet}")
    ws = wb[sheet]

    spec = SPECS[mode]
    if mode == "exogeneity":
        check_exogeneity(ws, expected_rows, marker, path.name, sheet, spec)
        return

    row3 = spec["row3"]
    n_cols = len(row3)
    last_col = get_column_letter(n_cols)
    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

    if ws["A1"].value in (None, ""):
        fail("missing title in A1")
    expected_merge = f"A1:{last_col}1"
    if expected_merge not in merged_ranges:
        fail(f"title row is not merged across {expected_merge}")
    if not ws["A1"].font.bold:
        fail("title is not bold")

    for col, expected in spec["row2"].items():
        actual = ws.cell(2, col).value
        if actual != expected:
            fail(f"row 2 col {col} mismatch: {actual!r}")
        if not ws.cell(2, col).font.bold:
            fail(f"row 2 col {col} is not bold")
    for expected_range in spec["merges"]:
        if expected_range not in merged_ranges:
            fail(f"missing row-2 merge: {expected_range}")

    actual_headers = [
        ws.cell(3, col).value or "" for col in range(1, n_cols + 1)
    ]
    if actual_headers != row3:
        fail(f"header mismatch: {actual_headers!r}")
    for col in range(2, n_cols + 1):
        cell = ws.cell(3, col)
        if not cell.font.bold:
            fail(f"header col {col} is not bold")
        if cell.border.bottom.style is None:
            fail(f"header col {col} has no bottom border")
        width = ws.column_dimensions[get_column_letter(col)].width
        if width is None or width < 8:
            fail(f"column {col} width is too small")

    if expected_rows is not None:
        note_row = 4 + expected_rows
        if ws.cell(note_row, 2).value in (None, ""):
            fail(f"missing footnote at row {note_row}")
        expected_note_merge = f"B{note_row}:{last_col}{note_row}"
        if expected_note_merge not in merged_ranges:
            fail(f"footnote row is not merged across {expected_note_merge}")
        if not ws.cell(note_row, 2).font.italic:
            fail(f"footnote row {note_row} is not italic")

    divider = spec.get("divider")
    if divider is not None:
        drow = divider["row"]
        if ws.cell(drow, 3).value != divider["label"]:
            fail(f"divider row {drow} label mismatch: {ws.cell(drow, 3).value!r}")
        if not ws.cell(drow, 3).font.bold:
            fail(f"divider row {drow} is not bold")
        # House style: the divider is bracketed by full-width horizontal rules
        # and is NOT merged; diagnostic values sit plainly in column C.
        divider_merge = f"C{drow}:{last_col}{drow}"
        if divider_merge in merged_ranges:
            fail(f"divider row should not be merged: {divider_merge}")
        if ws.cell(drow, 2).border.top.style is None:
            fail(f"divider row {drow} has no top border")
        if ws.cell(drow, 2).border.bottom.style is None:
            fail(f"divider row {drow} has no bottom border")
        if expected_rows is not None:
            # Value rows run from the divider+1 to just above the footnote.
            note_row = 4 + expected_rows
            for vrow in range(drow + 1, note_row):
                value_merge = f"C{vrow}:{last_col}{vrow}"
                if value_merge in merged_ranges:
                    fail(f"value row should not be merged: {value_merge}")

    string_probe = ws[spec["probe"]]
    if string_probe.value not in (None, "") and not isinstance(string_probe.value, str):
        fail(f"rendered cell is not stored as text: {string_probe.value!r}")

    if marker is not None:
        marker.write_text("ok\n", encoding="utf-8")
    print(f"PASS: {path.name}:{sheet} styled {mode} worksheet")


def check_exogeneity(
    ws,
    expected_rows: int | None,
    marker: Path | None,
    workbook_name: str,
    sheet: str,
    spec: dict,
) -> None:
    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}

    group_starts = []
    for col in range(3, ws.max_column + 1, 3):
        if ws.cell(2, col).value not in (None, ""):
            group_starts.append(col)

    if not group_starts:
        fail("no exogeneity group headers found")

    n_cols = group_starts[-1] + 2
    last_col = get_column_letter(n_cols)
    expected_title_merge = f"A1:{last_col}1"
    if expected_title_merge not in merged_ranges:
        fail(f"title row is not merged across {expected_title_merge}")
    if ws["A1"].value in (None, ""):
        fail("missing title in A1")
    if not ws["A1"].font.bold:
        fail("title is not bold")

    for start in group_starts:
        group_end = start + 2
        expected_merge = (
            f"{get_column_letter(start)}2:{get_column_letter(group_end)}2"
        )
        if expected_merge not in merged_ranges:
            fail(f"missing group-header merge: {expected_merge}")
        if not ws.cell(2, start).font.bold:
            fail(f"group header col {start} is not bold")
        actual_headers = [
            ws.cell(3, start + offset).value or "" for offset in range(3)
        ]
        if (
            actual_headers[0] != "HR"
            or not str(actual_headers[1]).endswith("% CI")
            or actual_headers[2] != "p-value"
        ):
            fail(f"exogeneity header block mismatch at col {start}: {actual_headers!r}")
        for col in range(start, group_end + 1):
            cell = ws.cell(3, col)
            if not cell.font.bold:
                fail(f"exogeneity header col {col} is not bold")
            if cell.border.bottom.style is None:
                fail(f"exogeneity header col {col} has no bottom border")
            width = ws.column_dimensions[get_column_letter(col)].width
            if width is None or width < 7:
                fail(f"column {col} width is too small")

    if ws["B4"].value in (None, ""):
        fail("first exogeneity term label is missing")

    joint_row = None
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 2).value == "Joint test (all lagged predictors)":
            joint_row = row
            break
    if joint_row is None:
        fail("missing joint-test row")

    if expected_rows is not None:
        expected_joint_row = 3 + expected_rows
        if joint_row != expected_joint_row:
            fail(f"joint-test row mismatch: {joint_row} != {expected_joint_row}")
        note_row = 4 + expected_rows
        if ws.cell(note_row, 2).value in (None, ""):
            fail(f"missing footnote at row {note_row}")
        expected_note_merge = f"B{note_row}:{last_col}{note_row}"
        if expected_note_merge not in merged_ranges:
            fail(f"footnote row is not merged across {expected_note_merge}")
        if not ws.cell(note_row, 2).font.italic:
            fail(f"footnote row {note_row} is not italic")

    string_probe = ws[spec["probe"]]
    if string_probe.value not in (None, "") and not isinstance(string_probe.value, str):
        fail(f"rendered cell is not stored as text: {string_probe.value!r}")

    if marker is not None:
        marker.write_text("ok\n", encoding="utf-8")
    print(f"PASS: {workbook_name}:{sheet} styled exogeneity worksheet")


def main(argv: list[str]) -> int:
    if len(argv) not in (4, 5, 6):
        print(
            "usage: check_iivw_xlsx.py WORKBOOK SHEET {balance|diagnostics|exogeneity} [expected_rows] [marker]",
            file=sys.stderr,
        )
        return 2
    expected_rows = int(argv[4]) if len(argv) >= 5 else None
    marker = Path(argv[5]) if len(argv) == 6 else None
    check(Path(argv[1]), argv[2], argv[3], expected_rows, marker)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
