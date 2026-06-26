#!/usr/bin/env python3
"""Assert the border/shade/zebra styling of an iivw reporting worksheet.

usage: check_iivw_style.py WORKBOOK SHEET {thin|medium|academic} SHADE ZEBRA [marker]

  SHADE / ZEBRA are 0 or 1.

Layout assumptions (the tabtools layout used by every iivw reporting command):
  row 1        merged title
  row 2        group spanners
  row 3        column headers
  rows 4..     data
Column A is a width-1 gutter; the table proper starts at column B (=2).
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def fail(message: str) -> None:
    print(f"FAIL: {message}", file=sys.stderr)
    raise SystemExit(1)


def has_fill(cell) -> bool:
    return cell.fill is not None and cell.fill.patternType == "solid"


def blank(value) -> bool:
    return value is None or value == ""


def note_row(ws) -> int:
    row = ws.max_row
    if row <= 3 or blank(ws.cell(row, 2).value):
        return 0
    for col in range(3, ws.max_column + 1):
        if not blank(ws.cell(row, col).value):
            return 0
    return row


def check(path: Path, sheet: str, mode: str, shade: bool, zebra: bool,
          marker: Path | None) -> None:
    if mode not in ("thin", "medium", "academic"):
        fail("border mode must be thin, medium, or academic")
    if not path.exists():
        fail(f"workbook not found: {path}")
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        fail(f"sheet not found: {sheet}")
    ws = wb[sheet]

    first_data = 4
    foot = note_row(ws)
    last_data = foot - 1 if foot else ws.max_row
    if last_data < first_data:
        fail("worksheet has no data rows to inspect")

    header = ws.cell(3, 2)        # B3
    data = ws.cell(first_data, 3)
    interior = ws.cell(first_data, 3) if first_data < last_data else None

    # Header always carries a bottom rule in every supported scheme.
    if header.border.bottom.style is None:
        fail("header row has no bottom border")

    if mode in ("thin", "medium"):
        expected = mode
        # House style (matches the regtab Tables 1-3 look): an outer frame plus
        # vertical separators after the label column and at each 3-column group
        # edge, with horizontal rules only in the header band.  NOT a full
        # interior grid.
        label = ws.cell(first_data, 2)            # B, label column
        right_edge = ws.cell(first_data, ws.max_column)
        if label.border.left.style != expected:
            fail(f"label cell B{first_data} left border is {label.border.left.style!r},"
                 f" expected {expected!r} (frame)")
        if label.border.right.style != expected:
            fail(f"label cell B{first_data} right border is {label.border.right.style!r},"
                 f" expected {expected!r} (separator after label column)")
        if right_edge.border.right.style != expected:
            fail(f"right-edge data cell has no {expected!r} right border (frame)")
        if interior is not None and (
            interior.border.top.style is not None
            or interior.border.bottom.style is not None
        ):
            fail(f"interior data cell C{first_data} has a horizontal rule;"
                 " thin/medium must not draw a full interior grid")
    else:  # academic: horizontal rules only, no interior verticals
        for side in ("left", "right"):
            got = getattr(data.border, side).style
            if got is not None:
                fail(f"academic data cell C{first_data} has a {side} border ({got!r});"
                     " academic must have no vertical rules")
        if interior is not None and interior.border.bottom.style is not None:
            fail("academic interior data row has a horizontal rule;"
                 " academic rules only the header and last row")

    # Header shading.
    if shade and not has_fill(header):
        fail("headershade requested but header has no fill")
    if not shade and has_fill(header):
        fail("header is shaded but headershade was not requested")

    # Zebra striping: the second data row is the first shaded stripe.  One-row
    # diagnostic tables have no alternating row to shade.
    stripe_row = first_data + 1
    if stripe_row <= last_data:
        stripe = ws.cell(stripe_row, 3)
        if zebra and not has_fill(stripe):
            fail("zebra requested but no stripe fill on the alternating data row")
        if not zebra and not shade and has_fill(stripe):
            fail("data row is shaded but neither zebra nor headershade was requested")

    if marker is not None:
        marker.write_text("ok\n", encoding="utf-8")
    print(f"PASS: {path.name}:{sheet} style {mode} shade={int(shade)} zebra={int(zebra)}")


def main(argv: list[str]) -> int:
    if len(argv) not in (6, 7):
        print(
            "usage: check_iivw_style.py WORKBOOK SHEET "
            "{thin|medium|academic} SHADE ZEBRA [marker]",
            file=sys.stderr,
        )
        return 2
    marker = Path(argv[6]) if len(argv) == 7 else None
    check(Path(argv[1]), argv[2], argv[3], bool(int(argv[4])),
          bool(int(argv[5])), marker)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
