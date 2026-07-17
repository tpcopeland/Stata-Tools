#!/usr/bin/env python3
"""
check_codescan_artifacts.py

Known-answer validation helper for codescan QA output artifacts.

Modes:
  xlsx  Validate the exact workbook produced by validation_codescan_output.do
  svg   Validate the exact SVG graph produced by validation_codescan_output.do

Writes PASS or FAIL to the requested result file and prints diagnostics to stdout.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def write_result(path: str, ok: bool) -> None:
    Path(path).write_text("PASS" if ok else "FAIL", encoding="utf-8")


def used_rows(ws) -> int:
    for row in range(ws.max_row, 0, -1):
        for col in range(ws.max_column, 0, -1):
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip() != "":
                return row
    return 0


def used_cols(ws) -> int:
    for col in range(ws.max_column, 0, -1):
        for row in range(ws.max_row, 0, -1):
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip() != "":
                return col
    return 0


def approx_equal(a: float, b: float, tol: float = 1e-9) -> bool:
    return abs(float(a) - float(b)) <= tol


def wilson_ci(count: int, n: int, level: float = 0.95) -> tuple[float, float]:
    z = 1.959963984540054
    p_hat = count / n
    z2n = (z * z) / n
    denom = 1 + z2n
    center = (p_hat + z2n / 2) / denom
    margin = z * math.sqrt((p_hat * (1 - p_hat) + z2n / 4) / n) / denom
    return max(0.0, (center - margin) * 100), min(100.0, (center + margin) * 100)


def check_font_block(ws, max_row: int, max_col: int, errors: list[str]) -> None:
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.font.name != "Calibri":
                errors.append(
                    f"{ws.title} {cell.coordinate}: expected font Calibri, got {cell.font.name}"
                )
            if not approx_equal(cell.font.sz or 0, 11.0):
                errors.append(
                    f"{ws.title} {cell.coordinate}: expected font size 11, got {cell.font.sz}"
                )
            if bool(cell.font.bold):
                errors.append(
                    f"{ws.title} {cell.coordinate}: expected non-bold text"
                )


def check_xlsx(path: str, number_format: str = "0") -> list[str]:
    """Validate the exported workbook.

    number_format is the Excel format the numeric cells must carry. Stata's
    `export excel` propagates the variable's display format, so codescan's
    format() option is observable here: the default %9.1f lands as "0" and
    format(%9.2f) lands as "0.00". Passing the expected format in is what makes
    a non-default format() testable rather than assumed.
    """
    errors: list[str] = []
    wb = load_workbook(path, data_only=True)

    if wb.sheetnames != ["Sheet1", "cooccurrence"]:
        errors.append(f"Expected sheets ['Sheet1', 'cooccurrence'], got {wb.sheetnames}")
        return errors

    summary = wb["Sheet1"]
    cooc = wb["cooccurrence"]

    # 3.0.0 layout: condition, label, matches, total_hits, positive_units,
    # prevalence, ci_low, ci_high, pattern, exclusion.
    expected_headers = [
        "condition", "label", "matches", "total_hits", "positive_units",
        "prevalence", "ci_low", "ci_high", "pattern", "exclusion",
    ]
    ncols = len(expected_headers)
    # Column letters are derived from the header list rather than written out,
    # so adding a column moves every downstream check with it. Hardcoded letters
    # are how this checker came to assert prevalence in C after it had moved.
    col_of = {name: get_column_letter(i) for i, name in enumerate(expected_headers, start=1)}

    if used_rows(summary) != 4 or used_cols(summary) != ncols:
        errors.append(
            f"Sheet1 expected used range A1:{get_column_letter(ncols)}4, "
            f"got rows={used_rows(summary)} cols={used_cols(summary)}"
        )
    if used_rows(cooc) != 4 or used_cols(cooc) != 4:
        errors.append(
            f"cooccurrence expected used range A1:D4, got rows={used_rows(cooc)} cols={used_cols(cooc)}"
        )

    if list(summary.merged_cells.ranges):
        errors.append(f"Sheet1 expected no merged cells, got {list(summary.merged_cells.ranges)}")
    if list(cooc.merged_cells.ranges):
        errors.append(
            f"cooccurrence expected no merged cells, got {list(cooc.merged_cells.ranges)}"
        )

    check_font_block(summary, 4, ncols, errors)
    check_font_block(cooc, 4, 4, errors)

    actual_headers = [summary.cell(row=1, column=col).value for col in range(1, ncols + 1)]
    if actual_headers != expected_headers:
        errors.append(f"Sheet1 headers expected {expected_headers}, got {actual_headers}")
        return errors

    expected_rows = [
        ("dm2", 3, 75.0, "E11", ""),
        ("htn", 2, 50.0, "I10", ""),
        ("asthma", 1, 25.0, "J45", ""),
    ]
    n = 4
    for i, (condition, count, prevalence, pattern, exclusion) in enumerate(expected_rows, start=2):
        low, high = wilson_ci(count, n)

        def cell(name):
            return summary[f"{col_of[name]}{i}"]

        def expect(name, want, tol=None):
            got = cell(name).value
            ok = approx_equal(got, want, tol) if tol is not None else got == want
            if not ok:
                errors.append(f"Sheet1 {cell(name).coordinate} ({name}): expected {want!r}, got {got!r}")

        expect("condition", condition)
        # No label() in this call, so label falls back to the condition name.
        expect("label", condition)
        expect("matches", count)
        # This export is not in countmode, so there is no hit total: the cell is
        # empty, not a copy of the unit count.
        if cell("total_hits").value is not None:
            errors.append(
                f"Sheet1 {cell('total_hits').coordinate}: expected empty total_hits without "
                f"countmode, got {cell('total_hits').value!r}"
            )
        expect("positive_units", count)
        expect("prevalence", prevalence, 1e-9)
        expect("ci_low", low, 1e-9)
        expect("ci_high", high, 1e-9)
        expect("pattern", pattern)
        expect("exclusion", exclusion)

        # The count columns are integers and unaffected by format(); the
        # prevalence/CI columns carry the format() format.
        for name in ("matches", "positive_units"):
            if cell(name).number_format != "0":
                errors.append(
                    f"Sheet1 {cell(name).coordinate}: expected number format 0, "
                    f"got {cell(name).number_format}"
                )
        for name in ("prevalence", "ci_low", "ci_high"):
            if cell(name).number_format != number_format:
                errors.append(
                    f"Sheet1 {cell(name).coordinate}: expected number format {number_format}, "
                    f"got {cell(name).number_format}"
                )

    cooc_headers = ["condition", "dm2", "htn", "asthma"]
    actual_cooc_headers = [cooc.cell(row=1, column=col).value for col in range(1, 5)]
    if actual_cooc_headers != cooc_headers:
        errors.append(f"cooccurrence headers expected {cooc_headers}, got {actual_cooc_headers}")

    expected_cooc_rows = [
        ("dm2", [3, 1, 1]),
        ("htn", [1, 2, 0]),
        ("asthma", [1, 0, 1]),
    ]
    for i, (condition, values) in enumerate(expected_cooc_rows, start=2):
        if cooc[f"A{i}"].value != condition:
            errors.append(f"cooccurrence A{i}: expected {condition}, got {cooc[f'A{i}'].value}")
        for offset, expected in enumerate(values, start=2):
            cell = cooc.cell(row=i, column=offset)
            if cell.value != expected:
                errors.append(
                    f"cooccurrence {cell.coordinate}: expected {expected}, got {cell.value}"
                )
            if cell.number_format != "0":
                errors.append(
                    f"cooccurrence {cell.coordinate}: expected number format 0, got {cell.number_format}"
                )

    return errors


def check_svg(path: str) -> list[str]:
    errors: list[str] = []
    text = Path(path).read_text(encoding="utf-8")
    matches = re.findall(r'<text x="[^"]+" y="([0-9.]+)"[^>]*>([^<]+)</text>', text)
    positions: dict[str, list[float]] = {}
    for y, label in matches:
        positions.setdefault(label, []).append(float(y))

    required = [
        "Condition Prevalence",
        "Prevalence (%)",
        "dm2",
        "htn",
        "asthma",
        "75.0",
        "50.0",
        "25.0",
    ]
    for item in required:
        if item not in positions:
            errors.append(f"SVG missing text element: {item}")

    if errors:
        return errors

    y_dm2 = min(positions["dm2"])
    y_htn = min(positions["htn"])
    y_asthma = min(positions["asthma"])
    y_75 = min(positions["75.0"])
    y_50 = min(positions["50.0"])
    y_25 = min(positions["25.0"])

    if not (y_dm2 < y_htn < y_asthma):
        errors.append(
            f"Expected graph label order dm2 < htn < asthma by y position, got {y_dm2}, {y_htn}, {y_asthma}"
        )
    if not (y_75 < y_50 < y_25):
        errors.append(
            f"Expected bar-label order 75.0 < 50.0 < 25.0 by y position, got {y_75}, {y_50}, {y_25}"
        )
    # Bar labels and condition labels share the same row but may differ by
    # scheme-dependent offsets (e.g. s2color vs plotplainblind), so use a
    # generous tolerance for Y-alignment while still catching gross errors.
    y_tol = 15.0
    if not approx_equal(y_dm2, y_75, y_tol):
        errors.append(f"Expected dm2 label to align with 75.0 bar label, got {y_dm2} vs {y_75}")
    if not approx_equal(y_htn, y_50, y_tol):
        errors.append(f"Expected htn label to align with 50.0 bar label, got {y_htn} vs {y_50}")
    if not approx_equal(y_asthma, y_25, y_tol):
        errors.append(
            f"Expected asthma label to align with 25.0 bar label, got {y_asthma} vs {y_25}"
        )

    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=["xlsx", "svg"])
    parser.add_argument("artifact")
    parser.add_argument("result_file")
    parser.add_argument(
        "--number-format",
        default="0",
        help="Excel number format the prevalence/CI cells must carry "
        "(default 0, matching codescan's default %%9.1f; pass 0.00 for format(%%9.2f))",
    )
    args = parser.parse_args()

    if args.mode == "xlsx":
        errors = check_xlsx(args.artifact, args.number_format)
    else:
        errors = check_svg(args.artifact)

    ok = len(errors) == 0
    write_result(args.result_file, ok)
    if ok:
        print(f"PASS: {args.mode} artifact validated")
        return 0

    print(f"FAIL: {args.mode} artifact validation")
    for err in errors:
        print(f"  - {err}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
