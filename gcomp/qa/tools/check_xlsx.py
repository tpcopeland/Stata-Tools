#!/usr/bin/env python3
"""
Excel assertion validator for Stata-generated tables.

CANONICAL SOURCE: Stata development CLI xlsx tool.
This is the single source of truth. Per-package qa/tools/check_xlsx.py copies
are byte-identical, self-contained deployments of this file.

Runs assertion checks against .xlsx files and reports PASS/FAIL results.
Designed for automated testing: Stata test scripts call this via shell,
then read a result file to assert pass/fail.

Usage:
    python3 check_xlsx.py FILE [--sheet NAME] [checks...] [--result-file PATH]

Exit codes:
    0 = all checks passed
    1 = one or more checks failed
    2 = error (file not found, bad arguments)

Requirements:
    openpyxl>=3.1.0

Author: Timothy P Copeland, Karolinska Institutet
"""

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell import MergedCell
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl",
          file=sys.stderr)
    sys.exit(2)


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class CheckResult:
    """Result of a single check."""
    name: str
    passed: bool
    message: str
    detail: str = ""


# =============================================================================
# Helper Functions
# =============================================================================

def get_used_rows(ws: Worksheet) -> int:
    """Return the last row with actual content (scans backwards)."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(max_row, 0, -1):
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip():
                return r
    return 0


def get_used_cols(ws: Worksheet) -> int:
    """Return the last column with actual content (scans backwards)."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for c in range(max_col, 0, -1):
        for r in range(1, max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip():
                return c
    return 0


def freeze_pane_ref(ws: Worksheet) -> Optional[str]:
    """Return the worksheet freeze-pane reference, normalized to A1 notation."""
    pane = getattr(ws, "freeze_panes", None)
    if pane is None:
        return None
    coordinate = getattr(pane, "coordinate", None)
    if coordinate:
        return str(coordinate)
    pane_text = str(pane)
    return pane_text or None


def cell_value(ws: Worksheet, ref: str) -> Any:
    """Get value from a cell reference like 'A1' or 'B3'."""
    return ws[ref].value


def row_values(ws: Worksheet, row_num: int) -> list:
    """Get all non-None values from a row."""
    values = []
    for col in range(1, get_used_cols(ws) + 1):
        val = ws.cell(row=row_num, column=col).value
        if val is not None:
            values.append(str(val).strip())
    return values


def all_text_values(ws: Worksheet) -> list[str]:
    """Collect all string values from the sheet."""
    texts = []
    for row in range(1, get_used_rows(ws) + 1):
        for col in range(1, get_used_cols(ws) + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                texts.append(str(val))
    return texts


# =============================================================================
# Structure Checks
# =============================================================================

def check_min_rows(ws: Worksheet, n: int) -> CheckResult:
    actual = get_used_rows(ws)
    passed = actual >= n
    return CheckResult(
        name=f"Row count >= {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected >= {n}, actual: {actual}",
    )


def check_max_rows(ws: Worksheet, n: int) -> CheckResult:
    actual = get_used_rows(ws)
    passed = actual <= n
    return CheckResult(
        name=f"Row count <= {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected <= {n}, actual: {actual}",
    )


def check_min_cols(ws: Worksheet, n: int) -> CheckResult:
    actual = get_used_cols(ws)
    passed = actual >= n
    return CheckResult(
        name=f"Column count >= {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected >= {n}, actual: {actual}",
    )


def check_max_cols(ws: Worksheet, n: int) -> CheckResult:
    actual = get_used_cols(ws)
    passed = actual <= n
    return CheckResult(
        name=f"Column count <= {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected <= {n}, actual: {actual}",
    )


def check_exact_rows(ws: Worksheet, n: int) -> CheckResult:
    actual = get_used_rows(ws)
    passed = actual == n
    return CheckResult(
        name=f"Row count == {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected {n}, actual: {actual}",
    )


def check_exact_cols(ws: Worksheet, n: int) -> CheckResult:
    actual = get_used_cols(ws)
    passed = actual == n
    return CheckResult(
        name=f"Column count == {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected {n}, actual: {actual}",
    )


def check_no_empty_rows(ws: Worksheet) -> CheckResult:
    """Check that there are no fully empty rows within the used range."""
    max_row = get_used_rows(ws)
    max_col = get_used_cols(ws)
    empty_rows = []
    for r in range(1, max_row + 1):
        row_empty = True
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip():
                row_empty = False
                break
        if row_empty:
            empty_rows.append(r)
    passed = len(empty_rows) == 0
    return CheckResult(
        name="No empty rows",
        passed=passed,
        message="no empty rows found" if passed else f"empty rows: {empty_rows}",
    )


# =============================================================================
# Header Checks
# =============================================================================

def check_header_row(ws: Worksheet, row_num: int, expected: list[str]) -> CheckResult:
    """Check that row_num contains all expected values (any column)."""
    actual = row_values(ws, row_num)
    missing = [v for v in expected if v not in actual]
    passed = len(missing) == 0
    return CheckResult(
        name=f"Header row {row_num} contains {expected}",
        passed=passed,
        message="all found" if passed else f"missing: {missing}",
        detail=f"actual values: {actual}" if not passed else "",
    )


def check_header_exact(ws: Worksheet, row_num: int, expected: list[str]) -> CheckResult:
    """Check that row_num has exactly these non-empty values in order."""
    actual = row_values(ws, row_num)
    passed = actual == expected
    return CheckResult(
        name=f"Header row {row_num} exact match",
        passed=passed,
        message="exact match" if passed else f"expected: {expected}",
        detail=f"actual: {actual}" if not passed else "",
    )


# =============================================================================
# Cell Checks
# =============================================================================

def check_cell_value_exact(ws: Worksheet, ref: str, expected: str) -> CheckResult:
    """Check that cell equals expected value."""
    actual = ws[ref].value
    actual_str = str(actual).strip() if actual is not None else ""
    passed = actual_str == expected
    return CheckResult(
        name=f"Cell {ref} == \"{expected}\"",
        passed=passed,
        message="matched" if passed else f"expected \"{expected}\", got \"{actual_str}\"",
    )


def check_cells_not_empty(ws: Worksheet, refs: list[str]) -> CheckResult:
    """Check that listed cells are not empty."""
    empty = []
    for ref in refs:
        val = ws[ref].value
        if val is None or str(val).strip() == "":
            empty.append(ref)
    passed = len(empty) == 0
    return CheckResult(
        name=f"Cells not empty: {', '.join(refs)}",
        passed=passed,
        message="all non-empty" if passed else f"empty cells: {empty}",
    )


def check_col_not_empty(ws: Worksheet, col_letter: str) -> CheckResult:
    """Check that a column has no empty cells in the data range (row 2+)."""
    col_idx = column_index_from_string(col_letter)
    max_row = get_used_rows(ws)
    empty_rows = []
    for r in range(2, max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val is None or str(val).strip() == "":
            empty_rows.append(r)
    passed = len(empty_rows) == 0
    return CheckResult(
        name=f"Column {col_letter} not empty (rows 2-{max_row})",
        passed=passed,
        message="no empty cells" if passed else f"empty in rows: {empty_rows[:10]}",
    )


def check_cell_approx(ws: Worksheet, ref: str, expected: float,
                       tolerance: float) -> CheckResult:
    """Check that a numeric cell value is within tolerance of expected."""
    actual = ws[ref].value
    if actual is None:
        return CheckResult(
            name=f"Cell {ref} ~= {expected} (±{tolerance})",
            passed=False,
            message="cell is empty",
        )
    try:
        actual_num = float(actual)
    except (ValueError, TypeError):
        return CheckResult(
            name=f"Cell {ref} ~= {expected} (±{tolerance})",
            passed=False,
            message=f"not numeric: \"{actual}\"",
        )
    passed = abs(actual_num - expected) <= tolerance
    return CheckResult(
        name=f"Cell {ref} ~= {expected} (±{tolerance})",
        passed=passed,
        message=f"actual: {actual_num}" if passed
        else f"expected {expected} ±{tolerance}, got {actual_num}",
    )


def check_cell_contains(ws: Worksheet, ref: str, substring: str) -> CheckResult:
    """Check that a cell value contains a substring."""
    actual = ws[ref].value
    actual_str = str(actual) if actual is not None else ""
    passed = substring in actual_str
    return CheckResult(
        name=f"Cell {ref} contains \"{substring}\"",
        passed=passed,
        message="found" if passed else f"not found in \"{actual_str}\"",
    )


def check_cell_regex(ws: Worksheet, ref: str, pattern: str) -> CheckResult:
    """Check that a cell value matches a regex pattern."""
    actual = ws[ref].value
    actual_str = str(actual) if actual is not None else ""
    try:
        passed = bool(re.search(pattern, actual_str))
    except re.error as e:
        return CheckResult(
            name=f"Cell {ref} matches /{pattern}/",
            passed=False,
            message=f"invalid regex: {e}",
        )
    return CheckResult(
        name=f"Cell {ref} matches /{pattern}/",
        passed=passed,
        message="matched" if passed else f"no match in \"{actual_str}\"",
    )


# =============================================================================
# Sheet-Level Checks
# =============================================================================

def check_sheet_count(wb, n: int) -> CheckResult:
    """Check that the workbook has exactly N sheets."""
    actual = len(wb.sheetnames)
    passed = actual == n
    return CheckResult(
        name=f"Sheet count == {n}",
        passed=passed,
        message=f"actual: {actual}" if passed
        else f"expected {n}, actual: {actual} ({', '.join(wb.sheetnames)})",
    )


def check_sheet_names(wb, expected: list[str]) -> CheckResult:
    """Check that all expected sheet names exist in the workbook."""
    actual = set(wb.sheetnames)
    missing = [n for n in expected if n not in actual]
    passed = len(missing) == 0
    return CheckResult(
        name=f"Sheet names: {expected}",
        passed=passed,
        message="all found" if passed
        else f"missing: {missing} (have: {wb.sheetnames})",
    )


# =============================================================================
# Format Checks
# =============================================================================

def check_bold_rows(ws: Worksheet, row_nums: list[int]) -> CheckResult:
    """Check that specified rows have at least one bold cell."""
    max_col = get_used_cols(ws)
    not_bold = []
    for r in row_nums:
        has_bold = False
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.font and cell.font.bold:
                has_bold = True
                break
        if not has_bold:
            not_bold.append(r)
    passed = len(not_bold) == 0
    return CheckResult(
        name=f"Bold rows: {row_nums}",
        passed=passed,
        message="all bold" if passed else f"not bold: rows {not_bold}",
    )


def check_merged_rows(ws: Worksheet, row_nums: list[int]) -> CheckResult:
    """Check that specified rows have at least one merged cell region."""
    merged_row_set = set()
    for merged_range in ws.merged_cells.ranges:
        bounds = merged_range.bounds  # (min_col, min_row, max_col, max_row)
        for r in range(bounds[1], bounds[3] + 1):
            merged_row_set.add(r)
    not_merged = [r for r in row_nums if r not in merged_row_set]
    passed = len(not_merged) == 0
    return CheckResult(
        name=f"Merged rows: {row_nums}",
        passed=passed,
        message="all have merges" if passed else f"no merges: rows {not_merged}",
    )


def check_font(ws: Worksheet, expected_name: str) -> CheckResult:
    """Check that the primary font matches expected name.

    Only cells with content count: empty cells carry the workbook default
    font, which would otherwise outvote the styled table on sparse sheets
    (and disagree with --theme, which already counts non-empty cells only).
    """
    from collections import Counter
    font_counts: Counter[str] = Counter()
    for row in range(1, get_used_rows(ws) + 1):
        for col in range(1, get_used_cols(ws) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if _stringify(cell.value) == "":
                continue
            name = cell.font.name if cell.font else None
            if name:
                font_counts[name] += 1
    if not font_counts:
        return CheckResult(
            name=f"Font is {expected_name}",
            passed=False,
            message="no fonts detected",
        )
    primary = font_counts.most_common(1)[0][0]
    passed = primary == expected_name
    return CheckResult(
        name=f"Font is {expected_name}",
        passed=passed,
        message=f"primary font: {primary}" if passed else f"expected {expected_name}, got {primary}",
    )


def check_fontsize(ws: Worksheet, expected_size: float) -> CheckResult:
    """Check that the primary font size matches (non-empty cells only)."""
    from collections import Counter
    size_counts: Counter[float] = Counter()
    for row in range(1, get_used_rows(ws) + 1):
        for col in range(1, get_used_cols(ws) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if _stringify(cell.value) == "":
                continue
            size = cell.font.size if cell.font else None
            if size:
                size_counts[size] += 1
    if not size_counts:
        return CheckResult(
            name=f"Font size is {expected_size}",
            passed=False,
            message="no font sizes detected",
        )
    primary = size_counts.most_common(1)[0][0]
    passed = abs(primary - expected_size) < 0.5
    return CheckResult(
        name=f"Font size is {expected_size}",
        passed=passed,
        message=f"primary size: {primary}" if passed else f"expected {expected_size}, got {primary}",
    )


def check_has_borders(ws: Worksheet) -> CheckResult:
    """Check that the sheet has any border styling."""
    for row in range(1, get_used_rows(ws) + 1):
        for col in range(1, get_used_cols(ws) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            border = cell.border
            if border:
                for side in (border.top, border.bottom, border.left, border.right):
                    if side and side.style:
                        return CheckResult(
                            name="Has borders",
                            passed=True,
                            message="borders detected",
                        )
    return CheckResult(
        name="Has borders",
        passed=False,
        message="no borders detected",
    )


def check_border_style(ws: Worksheet, expected_style: str) -> CheckResult:
    """Check that the primary border style matches."""
    from collections import Counter
    style_counts: Counter[str] = Counter()
    for row in range(1, get_used_rows(ws) + 1):
        for col in range(1, get_used_cols(ws) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            border = cell.border
            if border:
                for side in (border.top, border.bottom, border.left, border.right):
                    if side and side.style:
                        style_counts[side.style] += 1
    if not style_counts:
        return CheckResult(
            name=f"Border style is {expected_style}",
            passed=False,
            message="no borders detected",
        )
    primary = style_counts.most_common(1)[0][0]
    passed = primary == expected_style
    return CheckResult(
        name=f"Border style is {expected_style}",
        passed=passed,
        message=f"primary: {primary}" if passed else f"expected {expected_style}, got {primary}",
    )


# =============================================================================
# Pattern Checks
# =============================================================================

def _sheet_has_pattern(ws: Worksheet, pattern_name: str) -> bool:
    """Substring/regex content-pattern detection (per-package lineage semantics)."""
    texts = _all_text(ws)
    joined = "\n".join(texts)
    if pattern_name == "p-values":
        return any(re.search(r"(?i)(p-value|p\s*[=<]|<0\.\d+)", t) for t in texts)
    if pattern_name == "ci":
        return any(re.search(r"\([^)]*,\s*[^)]*\)", t) for t in texts) or "95% CI" in joined
    if pattern_name == "percentages":
        return any("%" in t for t in texts)
    if pattern_name == "mean-sd":
        return any(re.search(r"\d[^\n]*\([^,()%]*\)", t) for t in texts)
    if pattern_name == "n-equals":
        return any(re.search(r"\bN\s*=\s*\d+", t) for t in texts)
    if pattern_name == "sensitivity":
        return any("Sensitivity" in t for t in texts)
    if pattern_name == "rates":
        return any(re.search(r"\d+\.?\d*\s*\([\d\.\-]+-[\d\.]+\)", t) for t in texts)
    if pattern_name == "reference":
        return any(re.match(r"(?i)^(reference|ref\.?)$", t.strip()) for t in texts)
    return False


def check_patterns(ws: Worksheet, pattern_names: list[str]) -> CheckResult:
    """Check that the sheet contains specified content patterns."""
    missing = [n for n in pattern_names if not _sheet_has_pattern(ws, n)]
    passed = len(missing) == 0
    return CheckResult(
        name=f"Patterns: {', '.join(pattern_names)}",
        passed=passed,
        message="all found" if passed else f"missing: {missing}",
    )


# =============================================================================
# Styling Checks (per-package lineage: fills, italics, borders, widths, themes)
# =============================================================================
# These helpers and checks were unified from the per-package qa/tools copies so
# this module is the single source of truth. They intentionally use raw
# max_row/max_column scans (matching the original per-package behaviour the
# released test suites were written against).

def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # Python float repr is already the shortest form (no trailing zeros);
        # stripping characters would corrupt scientific notation (1e-10 ->
        # "1e-1").
        return str(value)
    return str(value).strip()


def _hex_to_rgb_string(hex_color: str) -> Optional[str]:
    if not hex_color:
        return None
    hex_color = hex_color.strip()
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    if len(hex_color) != 6:
        return None
    try:
        values = [str(int(hex_color[i:i + 2], 16)) for i in (0, 2, 4)]
    except ValueError:
        return None
    return " ".join(values)


def _cell_fill_rgb(cell) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "fill_type", None) != "solid":
        return None
    for color in (getattr(fill, "fgColor", None), getattr(fill, "start_color", None)):
        if color is None:
            continue
        rgb = getattr(color, "rgb", None)
        if rgb:
            parsed = _hex_to_rgb_string(rgb)
            if parsed is not None:
                return parsed
    return None


def _cell_border_matches(cell, side_name: str, style: str) -> bool:
    border = getattr(cell, "border", None)
    if border is None:
        return False
    side = getattr(border, side_name, None)
    return side is not None and getattr(side, "style", None) == style


def _iter_nonempty_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if _stringify(cell.value) != "":
                yield cell


def _all_text(ws) -> list[str]:
    return [_stringify(cell.value) for cell in _iter_nonempty_cells(ws)]


def _row_strings(ws, row: int) -> list[str]:
    return [_stringify(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)]


def _row_joined(ws, row: int) -> str:
    return " | ".join(v for v in _row_strings(ws, row) if v != "")


def _find_rows_containing(ws, text: str) -> list[int]:
    return [row for row in range(1, ws.max_row + 1) if text in _row_joined(ws, row)]


def _row_has_fill(ws, row: int, expected_rgb: Optional[str] = None) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        fill_rgb = _cell_fill_rgb(cell)
        if fill_rgb is None:
            continue
        if expected_rgb is None or fill_rgb == expected_rgb:
            return True
    return False


def _row_any_bold(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        if getattr(getattr(cell, "font", None), "bold", False):
            return True
    return False


def _row_all_bold(ws, row: int) -> bool:
    found = False
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        found = True
        if not getattr(getattr(cell, "font", None), "bold", False):
            return False
    return found


def _row_any_italic(ws, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if _stringify(cell.value) == "":
            continue
        if getattr(getattr(cell, "font", None), "italic", False):
            return True
    return False


def _dominant_font(ws: Worksheet) -> tuple[str, Optional[float]]:
    from collections import Counter
    counts: Counter[tuple[str, Optional[float]]] = Counter()
    for cell in _iter_nonempty_cells(ws):
        font = getattr(cell, "font", None)
        if font is None:
            continue
        name = (font.name or "").strip()
        size = round(float(font.sz), 1) if font.sz is not None else None
        counts[(name, size)] += 1
    if not counts:
        return ("", None)
    return counts.most_common(1)[0][0]


def _theme_expectations(name: str):
    theme = name.lower()
    if theme == "nejm":
        # regtab NEJM theme emits Arial 9pt (see _tabtools_common.ado: "9pt
        # matches NEJM manuscript table conventions (10pt was too large)").
        return ("Arial", 9.0, "medium")
    if theme == "lancet":
        return ("Arial", 9.0, "medium")
    if theme == "apa":
        return ("Times New Roman", 12.0, None)
    return None


def _column_width(ws, ref: str) -> float:
    ref = ref.strip().upper()
    try:
        letter = get_column_letter(int(ref))
    except ValueError:
        letter = get_column_letter(column_index_from_string(ref))
    width = ws.column_dimensions[letter].width
    return 8.43 if width is None else float(width)


def _merged_range_for_cell(ws, row: int, col_idx: int):
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row
                and merged.min_col <= col_idx <= merged.max_col):
            return merged
    return None


def _cell_effective_width(ws, row: int, col_idx: int) -> Optional[float]:
    merged = _merged_range_for_cell(ws, row, col_idx)
    if merged is not None:
        if row != merged.min_row or col_idx != merged.min_col:
            return None
        return sum(
            _column_width(ws, get_column_letter(i))
            for i in range(merged.min_col, merged.max_col + 1)
        )
    return _column_width(ws, get_column_letter(col_idx))


def _cell_width_fit_failures(ws, col_indices: list[int], start_row: int,
                             slack: float) -> list[dict[str, Any]]:
    failures = []
    for row in range(max(start_row, 1), ws.max_row + 1):
        for col_idx in col_indices:
            cell = ws.cell(row=row, column=col_idx)
            value = _stringify(cell.value)
            if not value:
                continue
            if getattr(getattr(cell, "alignment", None), "wrap_text", False):
                continue
            effective_width = _cell_effective_width(ws, row, col_idx)
            if effective_width is None:
                continue
            if effective_width + slack + 1e-9 < len(value):
                failures.append({
                    "col": get_column_letter(col_idx),
                    "width": round(effective_width, 2),
                    "max_len": len(value),
                    "row": row,
                    "sample": value[:40],
                })
    return failures


def check_sheet_order(wb, expected: list[str]) -> CheckResult:
    actual = wb.sheetnames
    passed = actual == expected
    return CheckResult(
        name=f"Sheet order {expected}",
        passed=passed,
        message="exact order" if passed else f"found: {actual}",
    )


def check_contains(ws: Worksheet, text: str) -> CheckResult:
    passed = any(text in t for t in _all_text(ws))
    return CheckResult(
        name=f"Workbook contains \"{text}\"",
        passed=passed,
        message="found" if passed else "not found",
    )


def check_row_contains(ws: Worksheet, row: int, text: str) -> CheckResult:
    row_text = _row_joined(ws, row)
    passed = text in row_text
    return CheckResult(
        name=f"Row {row} contains \"{text}\"",
        passed=passed,
        message="found" if passed else f"not in: {row_text}",
    )


def check_bold_row_all(ws: Worksheet, row: int) -> CheckResult:
    passed = _row_all_bold(ws, row)
    return CheckResult(
        name=f"Row {row} all bold",
        passed=passed,
        message="all bold" if passed else "not all non-empty cells bold",
    )


def check_italic_row(ws: Worksheet, row: int) -> CheckResult:
    passed = _row_any_italic(ws, row)
    return CheckResult(
        name=f"Row {row} italic",
        passed=passed,
        message="italic" if passed else "no italic cell",
    )


def check_italic_cell(ws: Worksheet, ref: str) -> CheckResult:
    passed = bool(getattr(getattr(ws[ref], "font", None), "italic", False))
    return CheckResult(
        name=f"Cell {ref} italic",
        passed=passed,
        message="italic" if passed else "not italic",
    )


def check_has_fill(ws: Worksheet, row: int) -> CheckResult:
    passed = _row_has_fill(ws, row)
    return CheckResult(
        name=f"Row {row} has fill",
        passed=passed,
        message="filled" if passed else "no solid fill",
    )


def check_fill_color(ws: Worksheet, row: int, rgb: str) -> CheckResult:
    passed = _row_has_fill(ws, row, rgb)
    return CheckResult(
        name=f"Row {row} fill {rgb}",
        passed=passed,
        message="matched" if passed else f"no fill '{rgb}'",
    )


def check_row_bold_contains(ws: Worksheet, text: str) -> CheckResult:
    rows = _find_rows_containing(ws, text)
    if not rows:
        return CheckResult(f"Bold row containing \"{text}\"", False, "no row contains text")
    passed = any(_row_any_bold(ws, r) for r in rows)
    return CheckResult(
        name=f"Bold row containing \"{text}\"",
        passed=passed,
        message="bold" if passed else "not bold",
    )


def check_row_fill_contains(ws: Worksheet, text: str, rgb: str) -> CheckResult:
    rows = _find_rows_containing(ws, text)
    if not rows:
        return CheckResult(f"Filled row containing \"{text}\"", False, "no row contains text")
    passed = any(_row_has_fill(ws, r, rgb) for r in rows)
    return CheckResult(
        name=f"Filled row containing \"{text}\"",
        passed=passed,
        message=f"fill {rgb}" if passed else f"no fill '{rgb}'",
    )


def check_border_row(ws: Worksheet, row: int, side: str, style: str) -> CheckResult:
    passed = any(
        _cell_border_matches(ws.cell(row=row, column=col), side, style)
        for col in range(1, ws.max_column + 1)
    )
    return CheckResult(
        name=f"Row {row} {style} {side} border",
        passed=passed,
        message="matched" if passed else f"no {style} {side} border",
    )


def check_cell_border(ws: Worksheet, ref: str, side: str, style: str) -> CheckResult:
    passed = _cell_border_matches(ws[ref], side, style)
    return CheckResult(
        name=f"Cell {ref} {style} {side} border",
        passed=passed,
        message="matched" if passed else f"no {style} {side} border on {ref}",
    )


def check_cell_no_fill(ws: Worksheet, ref: str) -> CheckResult:
    rgb = _cell_fill_rgb(ws[ref])
    passed = rgb is None
    return CheckResult(
        name=f"Cell {ref} has no fill",
        passed=passed,
        message="no fill" if passed else f"unexpected fill {rgb}",
    )


def check_cell_wrap(ws: Worksheet, ref: str) -> CheckResult:
    wrap_text = getattr(getattr(ws[ref], "alignment", None), "wrap_text", False)
    passed = bool(wrap_text)
    return CheckResult(
        name=f"Cell {ref} wraps text",
        passed=passed,
        message="wrap_text enabled" if passed else "wrap_text not enabled",
    )


def check_min_merges(ws: Worksheet, n: int) -> CheckResult:
    actual = len(ws.merged_cells.ranges)
    passed = actual >= n
    return CheckResult(
        name=f"Merged ranges >= {n}",
        passed=passed,
        message=f"actual: {actual}" if passed else f"expected >= {n}, actual: {actual}",
    )


def check_no_empty_cols(ws: Worksheet) -> CheckResult:
    max_row = get_used_rows(ws)
    max_col = get_used_cols(ws)
    for col in range(1, max_col + 1):
        if all(_stringify(ws.cell(row=r, column=col).value) == ""
               for r in range(1, max_row + 1)):
            return CheckResult("No empty columns", False, f"column {col} is empty")
    return CheckResult("No empty columns", True, "no empty columns")


def check_has_freeze_panes(ws: Worksheet) -> CheckResult:
    actual = freeze_pane_ref(ws)
    return CheckResult(
        name="Has freeze panes",
        passed=actual is not None,
        message=f"freeze panes at {actual}" if actual else "no freeze panes",
    )


def check_freeze_panes(ws: Worksheet, expected_ref: str) -> CheckResult:
    actual = freeze_pane_ref(ws)
    expected = expected_ref.strip().upper()
    actual_norm = actual.upper() if actual else None
    passed = actual_norm == expected
    return CheckResult(
        name=f"Freeze panes at {expected}",
        passed=passed,
        message=f"matched ({actual})" if passed else f"expected {expected}, got {actual or 'none'}",
    )


def check_number_format(ws: Worksheet, ref: str, fmt: str) -> CheckResult:
    actual = getattr(ws[ref], "number_format", None)
    passed = actual == fmt
    return CheckResult(
        name=f"Cell {ref} number format {fmt}",
        passed=passed,
        message="matched" if passed else f"expected '{fmt}', got '{actual}'",
    )


def check_col_width_at_least(ws: Worksheet, col: str, width: float) -> CheckResult:
    actual = _column_width(ws, col)
    passed = actual + 1e-9 >= width
    return CheckResult(
        name=f"Column {col} width >= {width}",
        passed=passed,
        message=f"actual: {actual:.2f}" if passed else f"width {actual:.2f} < {width:.2f}",
    )


def check_col_width_at_most(ws: Worksheet, col: str, width: float) -> CheckResult:
    actual = _column_width(ws, col)
    passed = actual - 1e-9 <= width
    return CheckResult(
        name=f"Column {col} width <= {width}",
        passed=passed,
        message=f"actual: {actual:.2f}" if passed else f"width {actual:.2f} > {width:.2f}",
    )


def check_col_width_fits_content(ws: Worksheet, col: str, start_row: int) -> CheckResult:
    try:
        col_idx = int(col.strip())
    except ValueError:
        col_idx = column_index_from_string(col.strip().upper())
    failures = _cell_width_fit_failures(ws, [col_idx], start_row, 0.0)
    if failures:
        failure = failures[0]
        message = (
            f"width {failure['width']:.2f} < content len {failure['max_len']} "
            f"(row {failure['row']}: '{failure['sample']}')"
        )
    else:
        message = "all unwrapped cells fit"
    return CheckResult(
        name=f"Column {col} width fits content",
        passed=not failures,
        message=message,
    )


def check_all_col_widths_fit(ws: Worksheet, start_row: int, slack: float) -> CheckResult:
    max_col = get_used_cols(ws)
    failures = _cell_width_fit_failures(
        ws,
        list(range(1, max_col + 1)),
        start_row,
        slack,
    )
    passed = not failures
    preview = ", ".join(
        f"{f['col']} width {f['width']:.2f} < len {f['max_len']} (row {f['row']})"
        for f in failures[:5]
    )
    more = f"; +{len(failures) - 5} more" if len(failures) > 5 else ""
    return CheckResult(
        name=f"All unwrapped column widths fit content from row {start_row} (slack {slack})",
        passed=passed,
        message="all fit" if passed else preview + more,
        detail=f"failures: {failures}" if failures else "",
    )


def check_theme(ws: Worksheet, name: str) -> CheckResult:
    expected = _theme_expectations(name)
    if expected is None:
        return CheckResult(f"Theme {name}", False, f"unknown theme '{name}'")
    exp_font, exp_size, exp_border = expected
    dom_name, dom_size = _dominant_font(ws)
    problems = []
    if dom_name.lower() != exp_font.lower():
        problems.append(f"font '{dom_name}' != '{exp_font}'")
    if dom_size is None or abs(dom_size - exp_size) > 0.1:
        problems.append(f"size {dom_size} != {exp_size}")
    if exp_border is not None and not any(
        _cell_border_matches(cell, "bottom", exp_border)
        or _cell_border_matches(cell, "top", exp_border)
        for cell in _iter_nonempty_cells(ws)
    ):
        problems.append(f"missing {exp_border} borders")
    passed = not problems
    return CheckResult(
        name=f"Theme {name}",
        passed=passed,
        message="matched" if passed else "; ".join(problems),
    )


# =============================================================================
# Check Runner
# =============================================================================

class CheckRunner:
    """Loads workbook, runs checks, collects results."""

    def __init__(self, filepath: str, sheet_name: Optional[str] = None):
        self.filepath = Path(filepath)
        self.sheet_name = sheet_name
        self.wb: Optional[Workbook] = None
        self.ws: Optional[Worksheet] = None
        self.results: list[CheckResult] = []

    def load(self) -> bool:
        """Load the workbook and select the sheet."""
        if not self.filepath.exists():
            print(f"Error: File not found: {self.filepath}", file=sys.stderr)
            return False
        try:
            self.wb = load_workbook(str(self.filepath), data_only=True)
            if self.sheet_name:
                if self.sheet_name in self.wb.sheetnames:
                    self.ws = self.wb[self.sheet_name]
                else:
                    print(f"Error: Sheet '{self.sheet_name}' not found. "
                          f"Available: {', '.join(self.wb.sheetnames)}",
                          file=sys.stderr)
                    return False
            else:
                active = self.wb.active
                if not isinstance(active, Worksheet):
                    print("Error: workbook has no active worksheet.", file=sys.stderr)
                    return False
                self.ws = active
                self.sheet_name = active.title
            return True
        except Exception as e:
            print(f"Error loading file: {e}", file=sys.stderr)
            return False

    def run(self, args: argparse.Namespace) -> list[CheckResult]:
        """Parse args into checks and run them all."""
        ws = self.ws
        wb = self.wb
        if wb is None or ws is None:
            raise RuntimeError("CheckRunner.run() called before a successful load()")
        results = []

        # Sheet-level checks (operate on workbook, not worksheet)
        if args.sheet_count is not None:
            results.append(check_sheet_count(wb, args.sheet_count))
        if args.sheet_names:
            results.append(check_sheet_names(wb, args.sheet_names))

        # Structure checks
        if args.min_rows is not None:
            results.append(check_min_rows(ws, args.min_rows))
        if args.max_rows is not None:
            results.append(check_max_rows(ws, args.max_rows))
        if args.exact_rows is not None:
            results.append(check_exact_rows(ws, args.exact_rows))
        if args.min_cols is not None:
            results.append(check_min_cols(ws, args.min_cols))
        if args.max_cols is not None:
            results.append(check_max_cols(ws, args.max_cols))
        if args.exact_cols is not None:
            results.append(check_exact_cols(ws, args.exact_cols))
        if args.no_empty_rows:
            results.append(check_no_empty_rows(ws))

        # Header checks
        if args.header_row:
            for spec in args.header_row:
                row_num = int(spec[0])
                expected = spec[1:]
                results.append(check_header_row(ws, row_num, expected))
        if args.header_exact:
            for spec in args.header_exact:
                row_num = int(spec[0])
                expected = spec[1:]
                results.append(check_header_exact(ws, row_num, expected))

        # Cell checks
        if args.cell:
            for pair in args.cell:
                ref = pair[0]
                value = pair[1]
                results.append(check_cell_value_exact(ws, ref, value))
        if args.cell_approx:
            for spec in args.cell_approx:
                ref = spec[0]
                value = float(spec[1])
                tol = float(spec[2])
                results.append(check_cell_approx(ws, ref, value, tol))
        if args.cell_contains:
            for spec in args.cell_contains:
                ref = spec[0]
                substring = spec[1]
                results.append(check_cell_contains(ws, ref, substring))
        if args.cell_regex:
            for spec in args.cell_regex:
                ref = spec[0]
                pattern = spec[1]
                results.append(check_cell_regex(ws, ref, pattern))
        if args.cell_not_empty:
            results.append(check_cells_not_empty(ws, args.cell_not_empty))
        if args.col_not_empty:
            for col in args.col_not_empty:
                results.append(check_col_not_empty(ws, col))

        # Format checks
        if args.bold_row:
            results.append(check_bold_rows(ws, args.bold_row))
        if args.merged_row:
            results.append(check_merged_rows(ws, args.merged_row))
        if args.font:
            results.append(check_font(ws, args.font))
        if args.fontsize is not None:
            results.append(check_fontsize(ws, args.fontsize))
        if args.has_borders:
            results.append(check_has_borders(ws))
        if args.border_style:
            results.append(check_border_style(ws, args.border_style))

        # Pattern checks
        if args.has_pattern:
            results.append(check_patterns(ws, args.has_pattern))

        # Styling checks (per-package lineage)
        if args.sheet_order:
            results.append(check_sheet_order(wb, args.sheet_order))
        if args.contains:
            for text in args.contains:
                results.append(check_contains(ws, text))
        if args.row_contains:
            for spec in args.row_contains:
                results.append(check_row_contains(ws, int(spec[0]), spec[1]))
        if args.bold_row_all:
            for r in args.bold_row_all:
                results.append(check_bold_row_all(ws, r))
        if args.italic_row:
            for r in args.italic_row:
                results.append(check_italic_row(ws, r))
        if args.italic_cell:
            for ref in args.italic_cell:
                results.append(check_italic_cell(ws, ref))
        if args.has_fill:
            for r in args.has_fill:
                results.append(check_has_fill(ws, r))
        if args.fill_color:
            for spec in args.fill_color:
                results.append(check_fill_color(ws, int(spec[0]), spec[1]))
        if args.border_row:
            for spec in args.border_row:
                results.append(check_border_row(ws, int(spec[0]), spec[1], spec[2]))
        if args.cell_border:
            for spec in args.cell_border:
                results.append(check_cell_border(ws, spec[0], spec[1], spec[2]))
        if args.cell_no_fill:
            for ref in args.cell_no_fill:
                results.append(check_cell_no_fill(ws, ref))
        if args.cell_wrap:
            for ref in args.cell_wrap:
                results.append(check_cell_wrap(ws, ref))
        if args.min_merges is not None:
            results.append(check_min_merges(ws, args.min_merges))
        if args.no_empty_cols:
            results.append(check_no_empty_cols(ws))
        if args.has_freeze_panes:
            results.append(check_has_freeze_panes(ws))
        if args.freeze_panes:
            results.append(check_freeze_panes(ws, args.freeze_panes))
        if args.number_format:
            for spec in args.number_format:
                results.append(check_number_format(ws, spec[0], spec[1]))
        if args.row_bold_contains:
            for text in args.row_bold_contains:
                results.append(check_row_bold_contains(ws, text))
        if args.row_fill_contains:
            for spec in args.row_fill_contains:
                results.append(check_row_fill_contains(ws, spec[0], spec[1]))
        if args.col_width_at_least:
            for spec in args.col_width_at_least:
                results.append(check_col_width_at_least(ws, spec[0], float(spec[1])))
        if args.col_width_at_most:
            for spec in args.col_width_at_most:
                results.append(check_col_width_at_most(ws, spec[0], float(spec[1])))
        if args.col_width_fits_content:
            for spec in args.col_width_fits_content:
                results.append(check_col_width_fits_content(ws, spec[0], int(spec[1])))
        if args.all_col_widths_fit:
            for spec in args.all_col_widths_fit:
                results.append(check_all_col_widths_fit(ws, int(spec[0]), float(spec[1])))
        if args.theme:
            for name in args.theme:
                results.append(check_theme(ws, name))

        self.results = results
        return results


# =============================================================================
# Output
# =============================================================================

def print_report(filepath: str, sheet_name: str, results: list[CheckResult],
                 quiet: bool = False, verbose: bool = False) -> None:
    """Print check results to stdout."""
    print(f"XLSX CHECK: {filepath} ({sheet_name})")

    for r in results:
        if quiet and r.passed:
            continue
        tag = "PASS" if r.passed else "FAIL"
        print(f"  [{tag}] {r.name} ({r.message})")
        if verbose and r.detail:
            print(f"         {r.detail}")

    passed = sum(1 for r in results if r.passed)
    total = len(results)
    status = "PASS" if passed == total else "FAIL"
    print(f"\nRESULT: {status} ({passed}/{total} passed)")


def write_result_file(path: str, results: list[CheckResult]) -> None:
    """Write PASS or FAIL to a result file for Stata integration."""
    all_passed = all(r.passed for r in results)
    Path(path).write_text("PASS" if all_passed else "FAIL")


# =============================================================================
# CLI
# =============================================================================

class HeaderAction(argparse.Action):
    """Custom action to collect multi-value header specifications.

    Accumulates into a list of lists: [[row, val1, val2, ...], ...]
    """
    def __call__(self, parser, namespace, values, option_string=None):
        items = getattr(namespace, self.dest, None) or []
        items.append(values)
        setattr(namespace, self.dest, items)


class CellPairAction(argparse.Action):
    """Custom action to collect cell REF VALUE pairs."""
    def __call__(self, parser, namespace, values, option_string=None):
        items = getattr(namespace, self.dest, None) or []
        items.append(values)
        setattr(namespace, self.dest, items)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Excel assertion validator for Stata-generated tables.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  python3 check_xlsx.py output.xlsx --min-rows 5 --min-cols 4
  python3 check_xlsx.py output.xlsx --exact-rows 15 --exact-cols 7
  python3 check_xlsx.py output.xlsx --bold-row 1 2 3 --has-borders
  python3 check_xlsx.py output.xlsx --cell A1 "Table Title" --font Arial
  python3 check_xlsx.py output.xlsx --cell-approx B4 0.045 0.001
  python3 check_xlsx.py output.xlsx --cell-contains A1 "Model"
  python3 check_xlsx.py output.xlsx --cell-regex B2 "^\\d+\\.\\d{2}$"
  python3 check_xlsx.py output.xlsx --sheet-count 3 --sheet-names Summary Details
  python3 check_xlsx.py output.xlsx --has-pattern p-values ci --result-file check.txt

Stata integration:
  ! python3 check_xlsx.py "`file'" --min-rows 5 --result-file "`testdir'/_check.txt"
  file open _fh using "`testdir'/_check.txt", read text
  file read _fh _line
  file close _fh
  assert "`_line'" == "PASS"

Available patterns for --has-pattern:
  p-values      Values like 0.0234 or <0.001
  ci            Confidence intervals like (1.23, 4.56)
  percentages   Values with %% or count (pct) format
  rates         Rate (CI) format like 12.3 (10.1-14.5)
  reference     "Reference" or "Ref." text
""",
    )

    parser.add_argument("file", help="Excel file to check (.xlsx)")
    parser.add_argument("--sheet", "-s",
                        help="Sheet name (default: active sheet)")

    # Sheet-level checks
    sheet_grp = parser.add_argument_group("Sheet-level checks")
    sheet_grp.add_argument("--sheet-count", type=int, metavar="N",
                           help="Assert exact number of sheets")
    sheet_grp.add_argument("--sheet-names", nargs="+", metavar="NAME",
                           help="Assert expected sheet names exist")

    # Structure checks
    struct = parser.add_argument_group("Structure checks")
    struct.add_argument("--min-rows", type=int, metavar="N",
                        help="Minimum row count")
    struct.add_argument("--max-rows", type=int, metavar="N",
                        help="Maximum row count")
    struct.add_argument("--exact-rows", type=int, metavar="N",
                        help="Exact row count")
    struct.add_argument("--min-cols", type=int, metavar="N",
                        help="Minimum column count")
    struct.add_argument("--max-cols", type=int, metavar="N",
                        help="Maximum column count")
    struct.add_argument("--exact-cols", type=int, metavar="N",
                        help="Exact column count")
    struct.add_argument("--no-empty-rows", action="store_true",
                        help="No fully empty rows in used range")

    # Header checks
    header = parser.add_argument_group("Header checks")
    header.add_argument("--header-row", nargs="+", action=HeaderAction,
                        metavar="N_TEXT",
                        help="Row N must contain listed values (e.g., --header-row 3 OR 95%% CI P-value)")
    header.add_argument("--header-exact", nargs="+", action=HeaderAction,
                        metavar="N_TEXT",
                        help="Row N must have exactly these values in order")

    # Cell checks
    cells = parser.add_argument_group("Cell checks")
    cells.add_argument("--cell", nargs=2, action=CellPairAction,
                       metavar=("REF", "VALUE"),
                       help="Cell must equal value (e.g., --cell A1 \"Title\")")
    cells.add_argument("--cell-approx", nargs=3, action=CellPairAction,
                       metavar=("REF", "VALUE", "TOL"),
                       help="Numeric cell within tolerance (e.g., --cell-approx B4 0.045 0.001)")
    cells.add_argument("--cell-contains", nargs=2, action=CellPairAction,
                       metavar=("REF", "STR"),
                       help="Cell contains substring (e.g., --cell-contains A1 \"Model\")")
    cells.add_argument("--cell-regex", nargs=2, action=CellPairAction,
                       metavar=("REF", "PATTERN"),
                       help="Cell matches regex (e.g., --cell-regex B2 \"^\\d+$\")")
    cells.add_argument("--cell-not-empty", nargs="+", metavar="REF",
                       help="Cell(s) must not be empty")
    cells.add_argument("--col-not-empty", nargs="+", metavar="COL",
                       help="Column has no empty cells in data range")

    # Format checks
    fmt = parser.add_argument_group("Format checks")
    fmt.add_argument("--bold-row", nargs="+", type=int, metavar="N",
                     help="Row(s) must have bold cells")
    fmt.add_argument("--merged-row", nargs="+", type=int, metavar="N",
                     help="Row(s) must have merged cells")
    fmt.add_argument("--font", metavar="NAME",
                     help="Primary font must be NAME")
    fmt.add_argument("--fontsize", type=float, metavar="N",
                     help="Primary font size must be N")
    fmt.add_argument("--has-borders", action="store_true",
                     help="Sheet must have borders")
    fmt.add_argument("--border-style", metavar="STYLE",
                     help="Primary border style (thin/medium)")

    # Pattern checks
    pats = parser.add_argument_group("Pattern checks")
    pats.add_argument("--has-pattern", nargs="+", metavar="NAME",
                      help="Content must contain pattern(s): p-values, ci, percentages, "
                           "mean-sd, n-equals, sensitivity, rates, reference")

    # Styling checks (per-package lineage: fills, italics, borders, widths, themes)
    style = parser.add_argument_group("Styling checks")
    style.add_argument("--sheet-order", "--exact-sheets", nargs="+", dest="sheet_order",
                       metavar="NAME", help="Workbook sheets in exactly this order")
    style.add_argument("--contains", action="append", metavar="TEXT",
                       help="Workbook contains text anywhere")
    style.add_argument("--row-contains", nargs=2, action=CellPairAction,
                       metavar=("N", "TEXT"), help="Row N contains text")
    style.add_argument("--bold-row-all", type=int, action="append", metavar="N",
                       help="All non-empty cells in row N are bold")
    style.add_argument("--italic-row", type=int, action="append", metavar="N",
                       help="Row N has an italic cell")
    style.add_argument("--italic-cell", action="append", metavar="REF",
                       help="Cell REF is italic")
    style.add_argument("--has-fill", type=int, action="append", metavar="N",
                       help="Row N has a solid fill")
    style.add_argument("--fill-color", nargs=2, action=CellPairAction,
                       metavar=("N", "RGB"), help="Row N has fill RGB (\"r g b\" decimals)")
    style.add_argument("--row-fill-contains", nargs=2, action=CellPairAction,
                       metavar=("TEXT", "RGB"), help="Rows containing TEXT have fill RGB")
    style.add_argument("--row-bold-contains", action="append", metavar="TEXT",
                       help="Rows containing TEXT are bold")
    style.add_argument("--border-row", nargs=3, action=CellPairAction,
                       metavar=("N", "SIDE", "STYLE"),
                       help="Row N has SIDE (top/bottom/left/right) border of STYLE")
    style.add_argument("--cell-border", nargs=3, action=CellPairAction,
                       metavar=("REF", "SIDE", "STYLE"),
                       help="Cell REF has SIDE (top/bottom/left/right) border of STYLE")
    style.add_argument("--cell-no-fill", nargs="+", metavar="REF",
                       help="Cell(s) REF have no solid fill")
    style.add_argument("--cell-wrap", nargs="+", metavar="REF",
                       help="Cell(s) REF have wrap_text enabled")
    style.add_argument("--min-merges", type=int, metavar="N",
                       help="At least N merged ranges")
    style.add_argument("--no-empty-cols", action="store_true",
                       help="No fully empty columns in used range")
    style.add_argument("--has-freeze-panes", action="store_true",
                       help="Sheet has freeze panes")
    style.add_argument("--freeze-panes", metavar="REF",
                       help="Sheet freeze panes exactly at REF, e.g. A3 or B2")
    style.add_argument("--number-format", nargs=2, action=CellPairAction,
                       metavar=("REF", "FORMAT"), help="Cell REF has number format FORMAT")
    style.add_argument("--col-width-at-least", nargs=2, action=CellPairAction,
                       metavar=("COL", "W"), help="Column COL width >= W")
    style.add_argument("--col-width-at-most", nargs=2, action=CellPairAction,
                       metavar=("COL", "W"), help="Column COL width <= W")
    style.add_argument("--col-width-fits-content", nargs=2, action=CellPairAction,
                       metavar=("COL", "STARTROW"),
                       help="Column COL width fits unwrapped content from STARTROW")
    style.add_argument("--all-col-widths-fit", nargs=2, action=CellPairAction,
                       metavar=("STARTROW", "SLACK"),
                       help="Every used unwrapped cell width plus SLACK fits content from STARTROW")
    style.add_argument("--theme", action="append", metavar="NAME",
                       help="Dominant font/size/borders match theme: nejm, lancet, apa")

    # Output options
    out = parser.add_argument_group("Output options")
    out.add_argument("--result-file", metavar="PATH",
                     help="Write PASS/FAIL to file for Stata integration")
    out.add_argument("--quiet", "-q", action="store_true",
                     help="Only show failures")
    out.add_argument("--verbose", "-v", action="store_true",
                     help="Show all check details")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    # Validate file extension
    if not args.file.lower().endswith(".xlsx"):
        print("Warning: File does not have .xlsx extension.", file=sys.stderr)

    # Load workbook
    runner = CheckRunner(args.file, args.sheet)
    if not runner.load():
        sys.exit(2)

    # Run checks
    results = runner.run(args)

    if not results:
        print(f"XLSX CHECK: {args.file} ({runner.sheet_name})")
        print("  No checks specified. Use --help to see available checks.")
        sys.exit(0)

    # Output
    print_report(args.file, runner.sheet_name, results,
                 quiet=args.quiet, verbose=args.verbose)

    if args.result_file:
        write_result_file(args.result_file, results)

    # Exit code
    all_passed = all(r.passed for r in results)
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
